#!/usr/bin/env python3
"""Export every public product in a Ruten store to a formatted Excel file."""

from __future__ import annotations

import argparse
import html as html_lib
import json
import logging
import math
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Iterable
from urllib.parse import urljoin, urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

if TYPE_CHECKING:
    from playwright.sync_api import Page, Response


DEFAULT_STORE_URL = "https://www.ruten.com.tw/store/qzecrvyn/"
DEFAULT_OUTPUT = f"ruten_products_{datetime.now():%Y%m%d}.xlsx"
RUTEN_HOSTS = {"www.ruten.com.tw", "ruten.com.tw", "goods.ruten.com.tw"}

TITLE_KEYS = (
    "title", "name", "item_name", "itemName", "product_name", "productName",
    "goods_name", "goodsName", "prod_name", "subject",
)
PRICE_KEYS = (
    "price", "current_price", "currentPrice", "sell_price", "sellPrice",
    "goods_price", "goodsPrice", "goods_sell_price", "goodsSellPrice",
    "direct_price", "directPrice", "priceRange", "amount",
)
QTY_KEYS = (
    "quantity", "qty", "stock", "stock_qty", "stockQty", "stock_quantity",
    "stockQuantity", "stockCount", "inventory", "inventoryLevel", "remain", "remaining",
    "available_quantity", "availableQuantity", "item_qty", "itemQty",
)
URL_KEYS = ("url", "link", "item_url", "itemUrl", "product_url", "productUrl")
ID_KEYS = (
    "id", "item_id", "itemId", "goods_id", "goodsId", "goods_no", "goodsNo",
    "goodsno", "gno", "product_id", "productId", "prod_id", "prodId",
)

PRICE_RE = re.compile(
    r"(?:NT\$|NTD|\u552e\u50f9|\u76f4\u8cfc\u50f9|\u50f9\u683c|\$)\s*[:\uff1a]?\s*([0-9][0-9,]*(?:\.[0-9]+)?)",
    re.IGNORECASE,
)
QTY_PATTERNS = (
    re.compile(r"(?:\u5eab\u5b58|\u5269\u9918|\u6578\u91cf|\u5c1a\u6709|\u5171)\s*[:\uff1a]?\s*([0-9][0-9,]*)\s*(?:\u4ef6|\u500b|\u7d44|\u5f35|\u4efd)?"),
    re.compile(r"([0-9][0-9,]*)\s*(?:\u4ef6|\u500b|\u7d44|\u5f35|\u4efd)\s*(?:\u53ef\u552e|\u73fe\u8ca8|\u8ca9\u552e)?"),
)
ITEM_ID_RE = re.compile(r"(?:item/show\?|item_id[=/])([A-Za-z0-9_-]{8,})", re.IGNORECASE)


@dataclass
class Product:
    name: str
    price: int | float | None = None
    quantity: int | None = None
    url: str = ""
    item_id: str = ""
    sources: set[str] = field(default_factory=set)

    @property
    def key(self) -> str:
        return self.item_id or canonical_url(self.url) or f"{self.name}|{self.price}"


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def canonical_url(url: str) -> str:
    if not url:
        return ""
    url = urljoin("https://www.ruten.com.tw/", url)
    match = re.search(r"https?://(?:www\.)?ruten\.com\.tw/item/show\?([A-Za-z0-9_-]+)", url)
    return f"https://www.ruten.com.tw/item/show?{match.group(1)}" if match else url.split("#", 1)[0]


def item_id_from_url(url: str) -> str:
    match = ITEM_ID_RE.search(url or "")
    return match.group(1) if match else ""


def parse_number(value: Any) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return None
        return int(value) if float(value).is_integer() else float(value)
    text = clean_text(value).replace(",", "")
    match = re.search(r"-?[0-9]+(?:\.[0-9]+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def parse_price(text: Any) -> int | float | None:
    if isinstance(text, (int, float)) and not isinstance(text, bool):
        number = parse_number(text)
    else:
        match = PRICE_RE.search(str(text or ""))
        number = parse_number(match.group(1)) if match else None
    if number is None or number < 0 or number > 100_000_000:
        return None
    return number


def parse_quantity(text: Any) -> int | None:
    if isinstance(text, (int, float)) and not isinstance(text, bool):
        number = parse_number(text)
        return int(number) if number is not None and number >= 0 else None
    raw = str(text or "")
    for pattern in QTY_PATTERNS:
        match = pattern.search(raw)
        if match:
            return int(match.group(1).replace(",", ""))
    return None


def first_value(data: dict[str, Any], keys: Iterable[str]) -> Any:
    for key in keys:
        if key in data and data[key] not in (None, "", [], {}):
            return data[key]
    return None


def nested_value(data: dict[str, Any], keys: Iterable[str]) -> Any:
    value = first_value(data, keys)
    if value not in (None, "", [], {}):
        return value
    for container_key in ("offer", "offers", "sale", "inventory", "sku", "product"):
        container = data.get(container_key)
        candidates = container if isinstance(container, list) else [container]
        for candidate in candidates:
            if isinstance(candidate, dict):
                value = nested_value(candidate, keys)
                if value not in (None, "", [], {}):
                    return value
    return None


def nested_numeric(data: dict[str, Any], keys: Iterable[str], quantity: bool = False) -> int | float | None:
    value = first_value(data, keys)
    parser = parse_quantity if quantity else parse_price
    parsed = parser(value)
    if parsed is not None:
        return parsed
    if isinstance(value, dict):
        for candidate in ("value", "amount", "price", "quantity", "stock", "count"):
            parsed = parser(value.get(candidate))
            if parsed is not None:
                return parsed
    if isinstance(value, list):
        for candidate in value:
            if isinstance(candidate, dict):
                parsed = nested_numeric(candidate, keys, quantity=quantity)
                if parsed is not None:
                    return parsed
    for container_key in ("offer", "offers", "sale", "inventory", "sku", "product"):
        container = data.get(container_key)
        if isinstance(container, dict):
            parsed = nested_numeric(container, keys, quantity=quantity)
            if parsed is not None:
                return parsed
    return None


def product_from_dict(data: dict[str, Any], source: str) -> Product | None:
    raw_name = first_value(data, TITLE_KEYS)
    name = clean_text(raw_name)
    if len(name) < 5 or len(name) > 500:
        return None

    raw_url = nested_value(data, URL_KEYS)
    raw_id = first_value(data, ID_KEYS)
    url = canonical_url(clean_text(raw_url)) if raw_url else ""
    item_id = item_id_from_url(url)
    if not item_id and raw_id is not None:
        candidate = clean_text(raw_id)
        if re.fullmatch(r"[A-Za-z0-9_-]{8,}", candidate):
            item_id = candidate
            url = f"https://www.ruten.com.tw/item/show?{candidate}"

    price = nested_numeric(data, PRICE_KEYS)
    quantity = nested_numeric(data, QTY_KEYS, quantity=True)
    if price is None or not (item_id or "/item/show?" in url):
        return None
    return Product(name=name, price=price, quantity=quantity, url=url, item_id=item_id, sources={source})


def products_from_json(payload: Any, source: str) -> list[Product]:
    found: list[Product] = []
    stack = [payload]
    visited = 0
    while stack and visited < 200_000:
        node = stack.pop()
        visited += 1
        if isinstance(node, dict):
            product = product_from_dict(node, source)
            if product:
                found.append(product)
            stack.extend(node.values())
        elif isinstance(node, list):
            stack.extend(node)
    return found


def merge_product(target: Product, incoming: Product) -> None:
    if (not target.name or len(incoming.name) > len(target.name)) and incoming.name:
        target.name = incoming.name
    if target.price is None and incoming.price is not None:
        target.price = incoming.price
    if target.quantity is None and incoming.quantity is not None:
        target.quantity = incoming.quantity
    if not target.url and incoming.url:
        target.url = incoming.url
    if not target.item_id and incoming.item_id:
        target.item_id = incoming.item_id
    target.sources.update(incoming.sources)


class ProductCollector:
    def __init__(self) -> None:
        self.products: dict[str, Product] = {}

    def add(self, product: Product) -> None:
        if not product.name or product.price is None:
            return
        key = product.key
        if key in self.products:
            merge_product(self.products[key], product)
        else:
            self.products[key] = product

    def add_many(self, products: Iterable[Product]) -> None:
        for product in products:
            self.add(product)

    def handle_response(self, response: Response) -> None:
        try:
            content_type = response.headers.get("content-type", "").lower()
            if response.status != 200 or "json" not in content_type:
                return
            lower_url = response.url.lower()
            if "ruten" not in lower_url:
                return
            payload = response.json()
            self.add_many(products_from_json(payload, f"JSON:{response.url}"))
        except Exception:
            return


DOM_SCRIPT = r"""
() => {
  const anchors = [...document.querySelectorAll('a[href*="/item/show?"]')];
  return anchors.map((a) => {
    let card = a;
    for (let i = 0; i < 8 && card.parentElement; i++) {
      const p = card.parentElement;
      const t = (p.innerText || '').trim();
      card = p;
      if (t.length >= 12 && /(?:NT\$|\$|\u552e\u50f9|\u76f4\u8cfc\u50f9|\u50f9\u683c)\s*[:\uff1a]?\s*[0-9,]+/.test(t)) break;
    }
    const img = a.querySelector('img');
    const title = a.getAttribute('title') || a.getAttribute('aria-label') ||
      (img && img.getAttribute('alt')) || (a.innerText || '').trim();
    const attrs = {};
    for (const el of [a, card]) {
      if (!el || !el.attributes) continue;
      for (const attr of el.attributes) {
        if (attr.name.startsWith('data-')) attrs[attr.name] = attr.value;
      }
    }
    return {
      url: a.href,
      title,
      text: (card.innerText || '').trim(),
      attrs,
    };
  });
}
"""


def products_from_dom(page: Page) -> list[Product]:
    products: list[Product] = []
    try:
        rows = page.evaluate(DOM_SCRIPT)
    except Exception:
        return products
    for row in rows:
        url = canonical_url(row.get("url", ""))
        item_id = item_id_from_url(url)
        text = row.get("text", "")
        attrs_text = " ".join(f"{key}={value}" for key, value in row.get("attrs", {}).items())
        name = clean_text(row.get("title"))
        if not name or name in {"\u5546\u54c1\u5716\u7247", "\u67e5\u770b\u5546\u54c1"}:
            lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
            name = next((line for line in lines if len(line) >= 5 and not PRICE_RE.search(line)), "")
        price = parse_price(f"{text} {attrs_text}")
        quantity = parse_quantity(f"{text} {attrs_text}")
        if item_id and name and price is not None:
            products.append(Product(name, price, quantity, url, item_id, {"DOM"}))
    return products


def contains_challenge(page: Page) -> bool:
    try:
        text = page.locator("body").inner_text(timeout=2_000).lower()
        return any(token in text for token in ("captcha", "\u9a57\u8b49\u78bc", "\u6a5f\u5668\u4eba\u9a57\u8b49", "\u5b58\u53d6\u906d\u62d2", "access denied"))
    except Exception:
        return False


def scroll_until_stable(page: Page, collector: ProductCollector, max_scrolls: int, delay: float) -> None:
    stable_rounds = 0
    previous = -1
    for _ in range(max_scrolls):
        collector.add_many(products_from_dom(page))
        current = len(collector.products)
        stable_rounds = stable_rounds + 1 if current == previous else 0
        if stable_rounds >= 4:
            break
        previous = current
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(max(250, int(delay * 1_000)))


def find_next_page(page: Page) -> Any:
    selectors = (
        'a[rel="next"]', 'button[aria-label*="\u4e0b\u4e00"]', 'a[aria-label*="\u4e0b\u4e00"]',
        'button:has-text("\u4e0b\u4e00\u9801")', 'a:has-text("\u4e0b\u4e00\u9801")',
        'button:has-text("\u4e0b\u4e00\u9875")', 'a:has-text("\u4e0b\u4e00\u9875")',
    )
    for selector in selectors:
        locator = page.locator(selector).last
        try:
            if locator.count() and locator.is_visible() and locator.is_enabled():
                disabled = locator.get_attribute("disabled") is not None
                aria_disabled = locator.get_attribute("aria-disabled") == "true"
                class_name = locator.get_attribute("class") or ""
                if not disabled and not aria_disabled and "disabled" not in class_name.lower():
                    return locator
        except Exception:
            continue
    return None


def scrape_store(
    store_url: str,
    show_browser: bool,
    max_pages: int,
    max_scrolls: int,
    delay: float,
) -> tuple[list[Product], dict[str, str], str]:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright

    collector = ProductCollector()
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=not show_browser)
        context = browser.new_context(
            locale="zh-TW",
            timezone_id="Asia/Taipei",
            viewport={"width": 1440, "height": 1000},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        page.on("response", collector.handle_response)
        list_url = store_list_url(store_url)
        logging.info("Opening the Ruten store product list: %s", list_url)
        page.goto(list_url, wait_until="domcontentloaded", timeout=90_000)
        page.wait_for_timeout(2_000)
        if contains_challenge(page):
            browser.close()
            raise RuntimeError(
                "Ruten displayed a verification or blocked-access page. "
                "Run again with --show-browser and complete verification manually."
            )

        visited_signatures: set[tuple[str, str]] = set()
        page_number = 0
        while True:
            page_number += 1
            before = len(collector.products)
            logging.info("Collecting page %d...", page_number)
            scroll_until_stable(page, collector, max_scrolls, delay)
            collector.add_many(products_from_dom(page))
            after = len(collector.products)
            logging.info("Collected %d products (%d new on this page).", after, after - before)

            urls = sorted(product.url for product in collector.products.values() if product.url)
            signature = (page.url, urls[-1] if urls else "")
            if signature in visited_signatures:
                logging.warning("A repeated page was detected. Pagination stopped.")
                break
            visited_signatures.add(signature)
            if max_pages and page_number >= max_pages:
                break
            next_page = find_next_page(page)
            if next_page is None:
                break
            try:
                next_page.scroll_into_view_if_needed()
                next_page.click()
                page.wait_for_load_state("domcontentloaded", timeout=45_000)
                page.wait_for_timeout(max(700, int(delay * 1_000)))
            except PlaywrightTimeoutError:
                logging.warning("The next page timed out. Keeping all products collected so far.")
                break
            except Exception as exc:
                logging.warning("Could not open the next page: %s", exc)
                break

        cookies = {cookie["name"]: cookie["value"] for cookie in context.cookies()}
        user_agent = page.evaluate("navigator.userAgent")
        browser.close()
    return list(collector.products.values()), cookies, user_agent


def detail_quantity_from_html(html: str, source_url: str) -> int | None:
    scripts = re.findall(
        r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    for script in scripts:
        try:
            payload = json.loads(html_lib.unescape(script).strip())
            products = products_from_json(payload, source_url)
            for product in products:
                if product.quantity is not None:
                    return product.quantity
            quantity = find_quantity_in_json(payload)
            if quantity is not None:
                return quantity
        except (json.JSONDecodeError, TypeError):
            continue

    patterns = (
        r'"(?:stockQuantity|stock_quantity|quantity|qty|remaining|remain)"\s*:\s*"?([0-9]+)',
        r'(?:\u5eab\u5b58|\u5269\u9918|\u6578\u91cf|\u5c1a\u6709)\s*[:\uff1a]?\s*([0-9][0-9,]*)\s*(?:\u4ef6|\u500b|\u7d44|\u5f35|\u4efd)?',
    )
    for pattern in patterns:
        match = re.search(pattern, html, re.IGNORECASE)
        if match:
            return int(match.group(1).replace(",", ""))
    visible_text = html_lib.unescape(re.sub(r"<[^>]+>", " ", html))
    return parse_quantity(visible_text)


def find_quantity_in_json(payload: Any) -> int | None:
    stack = [payload]
    while stack:
        node = stack.pop()
        if isinstance(node, dict):
            quantity = nested_numeric(node, QTY_KEYS, quantity=True)
            if quantity is not None:
                return int(quantity)
            stack.extend(node.values())
        elif isinstance(node, list):
            stack.extend(node)
    return None


def fetch_one_quantity(
    product: Product,
    cookies: dict[str, str],
    user_agent: str,
    delay: float,
) -> tuple[str, int | None, str | None]:
    import requests

    time.sleep(max(0, delay))
    try:
        response = requests.get(
            product.url,
            headers={
                "User-Agent": user_agent,
                "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.7",
                "Referer": DEFAULT_STORE_URL,
            },
            cookies=cookies,
            timeout=30,
        )
        response.raise_for_status()
        return product.key, detail_quantity_from_html(response.text, product.url), None
    except Exception as exc:
        return product.key, None, str(exc)


def fill_missing_quantities(
    products: list[Product],
    cookies: dict[str, str],
    user_agent: str,
    workers: int,
    delay: float,
    limit: int,
) -> None:
    missing = [product for product in products if product.quantity is None and product.url]
    if limit:
        missing = missing[:limit]
    if not missing:
        return
    logging.info(
        "Opening %d product pages to complete quantities (%d workers)...",
        len(missing),
        workers,
    )
    product_by_key = {product.key: product for product in products}
    completed = 0
    failures = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [
            executor.submit(fetch_one_quantity, product, cookies, user_agent, delay)
            for product in missing
        ]
        for future in as_completed(futures):
            key, quantity, error = future.result()
            completed += 1
            if quantity is not None:
                product_by_key[key].quantity = quantity
                product_by_key[key].sources.add("DETAIL")
            elif error:
                failures += 1
            if completed % 50 == 0 or completed == len(missing):
                logging.info(
                    "Quantity lookup progress: %d/%d (%d request failures).",
                    completed,
                    len(missing),
                    failures,
                )


def sort_products(products: list[Product]) -> list[Product]:
    return sorted(products, key=lambda product: (product.name.casefold(), product.item_id))


def export_excel(products: list[Product], output: Path, store_url: str) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.sheet_view.showGridLines = False
    headers = ["Product Name", "Quantity", "Price"]
    sheet.append(headers)

    for product in sort_products(products):
        sheet.append([product.name, product.quantity, product.price])
        name_cell = sheet.cell(sheet.max_row, 1)
        if product.url:
            name_cell.hyperlink = product.url
            name_cell.style = "Hyperlink"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    light_line = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="5B9BD5"))
    sheet.row_dimensions[1].height = 26

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).alignment = Alignment(vertical="center")
        sheet.cell(row, 2).alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(row, 3).alignment = Alignment(horizontal="right", vertical="center")
        sheet.cell(row, 2).number_format = "#,##0"
        sheet.cell(row, 3).number_format = '"NT$"#,##0'
        for col in range(1, 4):
            sheet.cell(row, col).border = Border(bottom=light_line)
        if row % 2 == 0:
            for col in range(1, 4):
                sheet.cell(row, col).fill = PatternFill("solid", fgColor="F7FAFC")

    sheet.column_dimensions["A"].width = 78
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 16
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:C{max(1, sheet.max_row)}"
    if sheet.max_row >= 2:
        table = Table(displayName="RutenProducts", ref=f"A1:C{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        sheet.add_table(table)

    info = workbook.create_sheet("Export Info")
    info.sheet_view.showGridLines = False
    missing_quantity = sum(product.quantity is None for product in products)
    info_rows = [
        ["Item", "Value"],
        ["Ruten Store", store_url],
        ["Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Product Count", len(products)],
        ["Unknown Quantity", missing_quantity],
        [
            "Notes",
            "Click a product name to open its Ruten page. Quantities that cannot be "
            "confirmed from public pages are left blank.",
        ],
    ]
    for row in info_rows:
        info.append(row)
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = header_font
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 90
    info["B2"].hyperlink = store_url
    info["B2"].style = "Hyperlink"
    info["B6"].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(output)


def validate_store_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or parsed.hostname not in RUTEN_HOSTS:
        raise argparse.ArgumentTypeError("Only Ruten store URLs under ruten.com.tw are accepted.")
    if "/store/" not in parsed.path:
        raise argparse.ArgumentTypeError(
            "The URL must be a Ruten store URL, for example "
            "https://www.ruten.com.tw/store/qzecrvyn/."
        )
    return url


def store_list_url(store_url: str) -> str:
    parsed = urlparse(store_url)
    match = re.search(r"/store/([^/]+)", parsed.path)
    if not match:
        return store_url
    seller = match.group(1)
    return f"{parsed.scheme}://{parsed.netloc}/store/{seller}/list"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export all public products from a Ruten store to Excel.")
    parser.add_argument("--store-url", type=validate_store_url, default=DEFAULT_STORE_URL, help="Ruten store URL")
    parser.add_argument("--output", type=Path, default=Path(DEFAULT_OUTPUT), help="Output .xlsx path")
    parser.add_argument("--show-browser", action="store_true", help="Show the browser for manual verification")
    parser.add_argument("--skip-detail-quantity", action="store_true", help="Skip per-product quantity lookups")
    parser.add_argument("--assume-quantity-one", action="store_true", help="Use 1 when quantity remains unknown")
    parser.add_argument("--max-pages", type=int, default=0, help="Maximum pages; 0 means all pages")
    parser.add_argument("--max-scrolls", type=int, default=80, help="Maximum automatic scrolls per page")
    parser.add_argument("--detail-limit", type=int, default=0, help="Maximum product pages to inspect; 0 means all")
    parser.add_argument("--detail-workers", type=int, default=4, choices=range(1, 9), metavar="1-8")
    parser.add_argument("--delay", type=float, default=0.6, help="Delay between requests and scrolls")
    parser.add_argument("--demo", action="store_true", help="Create one demo row without connecting to Ruten")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
    if args.output.suffix.lower() != ".xlsx":
        args.output = args.output.with_suffix(".xlsx")

    if args.demo:
        products = [
            Product(
                name="Demo Yu-Gi-Oh! card DBWS-JP023 (SR)",
                quantity=1,
                price=80,
                url=DEFAULT_STORE_URL,
                item_id="demo",
                sources={"DEMO"},
            )
        ]
        export_excel(products, args.output, args.store_url)
        logging.info("Demo workbook created: %s", args.output.resolve())
        return 0

    try:
        products, cookies, user_agent = scrape_store(
            args.store_url, args.show_browser, args.max_pages, args.max_scrolls, args.delay
        )
        if not products:
            raise RuntimeError(
                "No products were collected. Run again with --show-browser and check whether "
                "manual verification is required."
            )
        if not args.skip_detail_quantity:
            fill_missing_quantities(
                products, cookies, user_agent, args.detail_workers, args.delay, args.detail_limit
            )
        if args.assume_quantity_one:
            for product in products:
                if product.quantity is None:
                    product.quantity = 1
        export_excel(products, args.output, args.store_url)
        missing = sum(product.quantity is None for product in products)
        logging.info(
            "Completed: %s (%d products, %d quantities left blank).",
            args.output.resolve(),
            len(products),
            missing,
        )
        return 0
    except KeyboardInterrupt:
        logging.error("Stopped by the user.")
        return 130
    except Exception as exc:
        logging.error("Export failed: %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())

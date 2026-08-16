#!/usr/bin/env python3
"""Read a Ruten export workbook and list Yu-Gi-Oh! cards on Kapaipai.

The default action is a local preview only.  Actual listing requires --execute.
"""

from __future__ import annotations

import argparse
import contextlib
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import time
import unicodedata
import urllib.error
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
try:
    from playwright.sync_api import Locator, Page, TimeoutError as PlaywrightTimeoutError, sync_playwright
except ModuleNotFoundError:  # Preview/parser mode can still explain how to install Playwright.
    Locator = Page = Any  # type: ignore[misc,assignment]
    PlaywrightTimeoutError = TimeoutError  # type: ignore[assignment]
    sync_playwright = None  # type: ignore[assignment]


SEARCH_URL = "https://trade.kapaipai.tw/search"
DEFAULT_CONFIG = Path(__file__).with_name("kapaipai_rules.json")
HEADER_ALIASES = {
    "name": ("Product Name", "Name", "Title", "\u5546\u54c1\u540d\u7a31", "\u540d\u7a31", "\u5546\u54c1\u6a19\u984c", "\u6a19\u984c"),
    "quantity": ("Quantity", "Stock", "\u6578\u91cf", "\u5eab\u5b58", "\u4ef6\u6578"),
    "price": ("Price", "Sale Price", "\u50f9\u683c", "\u552e\u50f9", "\u5546\u54c1\u50f9\u683c"),
}
CARD_CODE_RE = re.compile(r"(?<![A-Z0-9])([A-Z0-9]{2,12})[-\u2010\u2011\u2012\u2013\u2014](JP\d{3,4})(?![A-Z0-9])", re.I)
CONSIGNOR_RE = re.compile(r"\u5bc4\u8ce3\s*[:\uff1a]\s*([A-Za-z0-9\u3400-\u9fff]+)", re.I)
PRICE_RE = re.compile(r"-?\d+(?:\.\d+)?")
RESULT_RARITY_RE = re.compile(r"^[A-Z][A-Z0-9]*(?:-[A-Z0-9]+)*$")
PARENTHESIZED_RE = re.compile(r"[\uff08(]([^()\uff08\uff09]*)[)\uff09]")
PROGRAM_VERSION = "2026.08.16-unified-10-category-state"
LOGIN_SETTLE_MS = 5000
SEARCH_UI_TIMEOUT_SECONDS = 30
SEARCH_MAX_ATTEMPTS = 3
GAME_FILTER_NAMES = {
    "\u5168\u90e8\u904a\u6232", "\u5bf6\u53ef\u5922\u65e5\u6587", "\u5bf6\u53ef\u5922\u7e41\u4e2d", "\u5bf6\u53ef\u5922\u82f1\u6587", "\u5bf6\u53ef\u5922\u7c21\u4e2d",
    "\u822a\u6d77\u738b\u65e5\u6587", "\u904a\u6232\u738b\u65e5\u6587", "UA\u65e5\u6587", "\u92fc\u5f48\u65e5\u6587", "VG\u65e5\u6587",
    "GA\u82f1\u6587", "\u6d1b\u5361\u7d0d\u82f1\u6587", "hololive\u65e5\u6587", "WS\u65e5\u6587", "WSR\u65e5\u6587", "WSB\u65e5\u6587",
    "\u6578\u78bc\u5bf6\u8c9d\u65e5\u6587", "\u7b26\u6587\u6230\u5834\u7e41\u4e2d", "\u7b26\u6587\u6230\u5834\u7c21\u4e2d", "\u54e5\u5409\u62c9", "\u67ef\u5357\u65e5\u6587",
    "\u9cf4\u6f6e\u65e5\u6587", "\u529b\u9738\u738b\u7e41\u4e2d", "\u9f8d\u73e0\u878d\u65e5\u6587", "\u78a7\u85cd\u822a\u7dda\u7e41\u4e2d", "\u8591\u9905\u4eba",
    "NivelArena", "\u95c7\u5f71\u8a69\u7ae0\u65e5\u6587", "\u6392\u7403\u5c11\u5e74\u65e5\u6587", "\u5e7b\u7378\u5e15\u9b6f\u65e5\u6587",
    "\u8d85\u82f1\u64ca\u6230\u82f1\u6587", "\u5468\u908a", "FIFA\u6536\u85cf",
}
@dataclass
class Listing:
    row: int
    name: str
    quantity: int
    price: int
    card_code: str
    rarity_hint: str
    rarity: str
    paper: str
    consignor: str
    note: str
    is_alt_art: bool = False
    source_url: str = ""
    status: str = "Pending"
    message: str = ""


@dataclass
class SearchResultCandidate:
    code_text: str
    rarity: str
    is_alt_art: bool
    locator: Any = None


class NeedsManualInput(RuntimeError):
    pass


class CardNotFound(NeedsManualInput):
    pass


class GameFilterError(NeedsManualInput):
    pass


class SubmissionUnverified(RuntimeError):
    pass


class TeeTextIO:
    """Mirror console output to a UTF-8 transcript without changing input()."""

    def __init__(self, console: Any, transcript: Any) -> None:
        self.console = console
        self.transcript = transcript

    def write(self, value: str) -> int:
        self.console.write(value)
        self.transcript.write(value)
        self.transcript.flush()
        return len(value)

    def flush(self) -> None:
        self.console.flush()
        self.transcript.flush()

    def isatty(self) -> bool:
        return bool(getattr(self.console, "isatty", lambda: False)())

    @property
    def encoding(self) -> str:
        return getattr(self.console, "encoding", "utf-8") or "utf-8"


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_for_match(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = re.sub(r"[\u200b-\u200f\u2060\ufeff]", "", text)
    return re.sub(r"\s+", "", text).casefold()


def normalize_code(value: str) -> str:
    match = CARD_CODE_RE.search(value.upper())
    return f"{match.group(1)}-{match.group(2)}" if match else ""


def parse_positive_int(value: Any, field: str, row: int) -> int:
    if value is None or clean(value) == "":
        raise ValueError(f"Row {row}: {field} is blank.")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
    else:
        match = PRICE_RE.search(clean(value).replace(",", ""))
        if not match:
            raise ValueError(f"Row {row}: {field} cannot be parsed: {value}")
        number = float(match.group())
    if number <= 0 or not number.is_integer():
        raise ValueError(f"Row {row}: {field} must be a positive integer: {value}")
    return int(number)


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    required = ("score", "game_filter", "rarity_aliases", "consignor_aliases")
    missing = [key for key in required if key not in config]
    if missing:
        raise ValueError(f"The rules file is missing required keys: {', '.join(missing)}")
    return config


def find_headers(sheet) -> tuple[int, dict[str, int]]:
    aliases = {key: {text.lower() for text in values} for key, values in HEADER_ALIASES.items()}
    for row_number in range(1, min(sheet.max_row, 20) + 1):
        found: dict[str, int] = {}
        for cell in sheet[row_number]:
            text = clean(cell.value).lower()
            for key, accepted in aliases.items():
                if text in accepted:
                    found[key] = cell.column
        if set(found) == set(HEADER_ALIASES):
            return row_number, found
    raise ValueError("The Product Name, Quantity, and Price columns were not found.")


def longest_alias_match(text: str, aliases: dict[str, str]) -> tuple[str, str]:
    normalized_text = normalize_for_match(text)
    candidates = [
        (clean(alias), code)
        for alias, code in aliases.items()
        if clean(alias) and normalize_for_match(alias) in normalized_text
    ]
    if not candidates:
        return "", ""
    alias, code = max(candidates, key=lambda item: len(item[0]))
    return alias, clean(code).upper()


def parenthesized_rarity_text(title: str) -> str:
    """Return only text enclosed by ASCII or full-width parentheses."""
    return " ".join(clean(match) for match in PARENTHESIZED_RE.findall(title) if clean(match))


def detect_alt_art(title: str) -> bool:
    """Treat \u7570\u5716 as a version marker only when it appears in parentheses."""
    return "\u7570\u5716" in normalize_for_match(parenthesized_rarity_text(title))


def detect_paper(title: str) -> str:
    if re.search(r"\u97d3\u7d19|\u97d3\u7248|\u97d3\u88fd", title, re.I):
        return ""
    if re.search(r"\u65e5\u7d19|\u65e5\u672c\u7d19|\u65e5\u88fd", title, re.I):
        return "\u65e5\u7d19"
    return ""


def detect_consignor(title: str, aliases: dict[str, str]) -> str:
    match = CONSIGNOR_RE.search(title)
    if not match:
        return ""
    raw = match.group(1).strip()
    for alias, display in aliases.items():
        if alias.lower() == raw.lower():
            return clean(display)
    return raw


def build_note(paper: str, score: str, consignor: str) -> str:
    parts = [part for part in (paper, score, f"\u5bc4\u8ce3:{consignor}" if consignor else "") if part]
    return "\uff5c".join(parts)


def parse_listing(row: int, name: str, quantity: Any, price: Any, url: str, config: dict[str, Any]) -> Listing:
    code = normalize_code(name)
    if not code:
        raise ValueError("A Yu-Gi-Oh! card code was not found.")
    rarity_text = parenthesized_rarity_text(name)
    hint, rarity = longest_alias_match(rarity_text, config["rarity_aliases"])
    paper = detect_paper(name)
    consignor = detect_consignor(name, config["consignor_aliases"])
    score = clean(config["score"])
    return Listing(
        row=row,
        name=clean(name),
        quantity=parse_positive_int(quantity, "Quantity", row),
        price=parse_positive_int(price, "Price", row),
        card_code=code,
        rarity_hint=hint,
        rarity=rarity,
        paper=paper,
        consignor=consignor,
        note=build_note(paper, score, consignor),
        is_alt_art=detect_alt_art(name),
        source_url=url,
    )


def read_listings(path: Path, config: dict[str, Any]) -> tuple[list[Listing], list[dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    if "Products" in workbook.sheetnames:
        sheet = workbook["Products"]
    elif "\u5546\u54c1\u6e05\u55ae" in workbook.sheetnames:
        sheet = workbook["\u5546\u54c1\u6e05\u55ae"]
    else:
        sheet = workbook.active
    header_row, columns = find_headers(sheet)
    listings: list[Listing] = []
    skipped: list[dict[str, Any]] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        name_cell = sheet.cell(row, columns["name"])
        name = clean(name_cell.value)
        if not name:
            continue
        if "\u904a\u6232\u738b" not in name and not normalize_code(name):
            skipped.append({"row": row, "name": name, "reason": "Not a recognizable Yu-Gi-Oh! product"})
            continue
        try:
            listing = parse_listing(
                row,
                name,
                sheet.cell(row, columns["quantity"]).value,
                sheet.cell(row, columns["price"]).value,
                name_cell.hyperlink.target if name_cell.hyperlink else "",
                config,
            )
            listings.append(listing)
        except ValueError as error:
            skipped.append({"row": row, "name": name, "reason": str(error)})
    workbook.close()
    return listings, skipped


def write_preview(source: Path, listings: list[Listing], skipped: list[dict[str, Any]]) -> Path:
    output = source.with_name(f"{source.stem}_kapaipai_preview.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Upload Preview"
    headers = [
        "Source Row",
        "Product Name",
        "Card Code",
        "Rarity Hint",
        "Kapaipai Rarity",
        "Alternate Art",
        "Quantity",
        "Price",
        "Paper",
        "Consignor",
        "Note",
        "Status",
    ]
    sheet.append(headers)
    for item in listings:
        status = "Ready" if item.rarity else "Resolve rarity during upload"
        sheet.append([
            item.row,
            item.name,
            item.card_code,
            item.rarity_hint,
            item.rarity,
            "Yes" if item.is_alt_art else "No",
            item.quantity,
            item.price,
            item.paper,
            item.consignor,
            item.note,
            status,
        ])
        if item.is_alt_art:
            alt_cell = sheet.cell(sheet.max_row, headers.index("Alternate Art") + 1)
            alt_cell.font = Font(bold=True, color="9C0006")
            alt_cell.fill = PatternFill("solid", fgColor="FFC7CE")
    skip_sheet = workbook.create_sheet("Skipped Items")
    skip_sheet.append(["Source Row", "Product Name", "Reason"])
    for item in skipped:
        skip_sheet.append([item["row"], item["name"], item["reason"]])
    for current in workbook.worksheets:
        for cell in current[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="333333")
            cell.alignment = Alignment(horizontal="center")
        current.freeze_panes = "A2"
        current.auto_filter.ref = current.dimensions
        for column in range(1, current.max_column + 1):
            width = max((len(clean(current.cell(row, column).value)) for row in range(1, current.max_row + 1)), default=8) + 2
            current.column_dimensions[get_column_letter(column)].width = min(max(width, 10), 60)
        for row in current.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(output)
    return output


def visible(locator: Locator) -> Locator | None:
    for index in range(locator.count()):
        candidate = locator.nth(index)
        try:
            if candidate.is_visible():
                return candidate
        except PlaywrightTimeoutError:
            pass
    return None


def click_visible_text(page: Page, text: str, exact: bool = True, timeout: int = 5000) -> None:
    candidate = visible(page.get_by_text(text, exact=exact))
    if candidate is None:
        candidate = visible(page.get_by_role("button", name=re.compile(re.escape(text))))
    if candidate is None:
        raise NeedsManualInput(f"The '{text}' button was not found.")
    candidate.click(timeout=timeout)


def is_logged_in(page: Page) -> bool:
    body = clean(page.locator("body").inner_text())
    if not body:
        return False
    normalized_body = normalize_for_match(body)
    return not (
        "/login" in page.url.lower()
        or "\u7528\u6236\u767b\u5165" in body
        or "justamoment" in normalized_body
        or "checkingyourbrowser" in normalized_body
        or "\u9a57\u8b49\u60a8\u662f\u771f\u4eba" in body
        or "\u57f7\u884c\u5b89\u5168\u6027\u9a57\u8b49" in body
        or visible(page.get_by_role("button", name=re.compile("\u7acb\u5373\u767b\u5165"))) is not None
    )


def stabilize_after_login(page: Page) -> None:
    print(f"Login confirmed. Waiting {LOGIN_SETTLE_MS // 1000} seconds for the page to stabilize.")
    page.wait_for_timeout(LOGIN_SETTLE_MS)
    if not page.url.startswith(SEARCH_URL):
        page.goto(SEARCH_URL, wait_until="domcontentloaded")
    wait_for_search_interface(page)


def wait_for_login(page: Page, timeout_seconds: int = 600) -> None:
    if is_logged_in(page):
        print("Reusing the existing login session.")
        stabilize_after_login(page)
        return
    print("Complete Cloudflare verification and Kapaipai login in the Chrome window.")
    print("The workflow will continue automatically after login.")
    deadline = time.monotonic() + timeout_seconds
    stable_checks = 0
    while time.monotonic() < deadline:
        page.wait_for_timeout(1000)
        if is_logged_in(page):
            stable_checks += 1
            if stable_checks >= 2:
                print("Login detected.")
                stabilize_after_login(page)
                return
        else:
            stable_checks = 0
    raise RuntimeError("Login was not completed within 10 minutes.")


def find_installed_browser(explicit: str = "") -> Path:
    candidates: list[Path] = []
    if explicit:
        candidates.append(Path(explicit.strip().strip('"')).expanduser())
    for env_name, relative in (
        ("LOCALAPPDATA", "Google/Chrome/Application/chrome.exe"),
        ("PROGRAMFILES", "Google/Chrome/Application/chrome.exe"),
        ("PROGRAMFILES(X86)", "Google/Chrome/Application/chrome.exe"),
        ("PROGRAMFILES", "Microsoft/Edge/Application/msedge.exe"),
        ("PROGRAMFILES(X86)", "Microsoft/Edge/Application/msedge.exe"),
    ):
        root = os.environ.get(env_name)
        if root:
            candidates.append(Path(root) / relative)
    for command in ("chrome.exe", "chrome", "msedge.exe", "microsoft-edge"):
        resolved = shutil.which(command)
        if resolved:
            candidates.append(Path(resolved))
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()
    if explicit:
        raise FileNotFoundError(f"The specified browser was not found: {explicit}")
    raise FileNotFoundError(
        "Google Chrome or Microsoft Edge was not found. Install Chrome or use "
        "--browser-path to specify chrome.exe."
    )


def free_local_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def wait_for_debug_browser(port: int, timeout: float = 30.0) -> None:
    endpoint = f"http://127.0.0.1:{port}/json/version"
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        try:
            with urllib.request.urlopen(endpoint, timeout=1) as response:
                if response.status == 200:
                    return
        except (OSError, urllib.error.URLError):
            time.sleep(0.25)
    raise RuntimeError(
        "Chrome started, but the local debugging connection could not be established. "
        "Close the Chrome window opened by this workflow and try again."
    )


def launch_manual_login_browser(source: Path, browser_path: str = "") -> tuple[subprocess.Popen[Any], int]:
    executable = find_installed_browser(browser_path)
    profile = source.parent / ".kapaipai_chrome_profile"
    port = free_local_port()
    command = [
        str(executable),
        f"--remote-debugging-port={port}",
        f"--user-data-dir={profile}",
        "--no-first-run",
        "--no-default-browser-check",
        "--new-window",
        SEARCH_URL,
    ]
    flags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
    process = subprocess.Popen(command, creationflags=flags)
    wait_for_debug_browser(port)
    print(f"\nBrowser opened: {executable.name}")
    return process, port


def select_kapaipai_page(context: Any) -> Page:
    for page in reversed(context.pages):
        if "trade.kapaipai.tw" in page.url:
            return page
    return context.new_page()


def close_update_dialog(page: Page) -> None:
    button = visible(page.get_by_role("button", name="\u6211\u77e5\u9053\u4e86", exact=True))
    if button:
        try:
            button.click(timeout=3000)
            page.wait_for_timeout(300)
            print("The feature-update dialog was closed automatically.")
        except PlaywrightTimeoutError:
            pass


def custom_game_filter_trigger(page: Page) -> tuple[Locator | None, str]:
    candidates = page.locator("div[style*='cursor: pointer']")
    for index in range(candidates.count()):
        candidate = candidates.nth(index)
        try:
            if not candidate.is_visible():
                continue
            if candidate.locator("xpath=ancestor::*[@role='dialog']").count():
                continue
            label = clean(candidate.inner_text())
            if label in GAME_FILTER_NAMES:
                return candidate, label
        except PlaywrightTimeoutError:
            continue
    return None, ""


def visible_game_filter_dialog(page: Page) -> Locator | None:
    dialogs = page.get_by_role("dialog")
    for index in range(dialogs.count()):
        dialog = dialogs.nth(index)
        try:
            if dialog.is_visible() and "\u904a\u6232\u7be9\u9078" in clean(dialog.inner_text()):
                return dialog
        except PlaywrightTimeoutError:
            continue
    return None


def current_game_filter(page: Page) -> str:
    _, custom_label = custom_game_filter_trigger(page)
    if custom_label:
        return custom_label
    for selector in ("[role=option][aria-selected=true]", ".ant-select-selection-item"):
        candidates = page.locator(selector)
        for index in range(candidates.count()):
            candidate = candidates.nth(index)
            try:
                label = clean(candidate.inner_text())
                if label in GAME_FILTER_NAMES:
                    return label
            except PlaywrightTimeoutError:
                continue
    return ""


def dismiss_game_filter_dialog(page: Page) -> None:
    if visible_game_filter_dialog(page) is None:
        return
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(350)
    except PlaywrightTimeoutError:
        pass


def click_filter_target(target: Locator) -> None:
    try:
        target.click(timeout=3000)
    except PlaywrightTimeoutError:
        target.click(timeout=3000, force=True)


def wait_for_search_interface(page: Page, timeout_seconds: int = SEARCH_UI_TIMEOUT_SECONDS) -> None:
    """Wait for the delayed SPA controls and dismiss update dialogs that appear late."""
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        close_update_dialog(page)
        custom_trigger, _ = custom_game_filter_trigger(page)
        combo = visible(page.get_by_role("combobox"))
        if custom_trigger is not None or combo is not None:
            return
        page.wait_for_timeout(250)
    raise GameFilterError(f"The game filter did not load within {timeout_seconds} seconds after login.")


def ensure_ygo_filter(
    page: Page,
    game_filter: str,
    max_attempts: int = 3,
    force_reselect: bool = False,
) -> None:
    wait_for_search_interface(page)
    last_error = ""
    for attempt in range(1, max_attempts + 1):
        close_update_dialog(page)
        custom_trigger, current_filter = custom_game_filter_trigger(page)
        dialog = visible_game_filter_dialog(page)
        if current_filter == game_filter and dialog is None and not force_reselect:
            return
        try:
            used_custom_dialog = custom_trigger is not None or dialog is not None
            if used_custom_dialog:
                if dialog is None:
                    if custom_trigger is None:
                        raise GameFilterError("The game-filter button was not found on the results page.")
                    click_filter_target(custom_trigger)
                    page.wait_for_timeout(600)
                    dialog = visible_game_filter_dialog(page)
                if dialog is None:
                    raise GameFilterError("The game-filter dialog could not be opened.")
                target = visible(dialog.get_by_text(game_filter, exact=True))
                if target is None:
                    raise GameFilterError(f"'{game_filter}' was not found in the game-filter dialog.")
                click_filter_target(target)
            else:
                combo = visible(page.get_by_role("combobox"))
                if combo is None:
                    raise GameFilterError("The game filter was not found.")
                click_filter_target(combo)
                page.wait_for_timeout(600)
                target = visible(page.get_by_role("option", name=game_filter, exact=True))
                if target is None:
                    target = visible(page.locator("[role=option]").filter(has_text=game_filter))
                if target is None:
                    raise GameFilterError(f"'{game_filter}' was not found in the game filter.")
                click_filter_target(target)

            for _ in range(20):
                page.wait_for_timeout(250)
                selected_filter = current_game_filter(page)
                open_dialog = visible_game_filter_dialog(page)
                if selected_filter == game_filter:
                    if used_custom_dialog and open_dialog is not None:
                        dismiss_game_filter_dialog(page)
                        open_dialog = visible_game_filter_dialog(page)
                    if open_dialog is None:
                        page.wait_for_timeout(500)
                        if current_game_filter(page) != game_filter:
                            last_error = "The category changed again before the search could start."
                            continue
                        print(f"Game category selected automatically: {game_filter}")
                        return
            selected_filter = current_game_filter(page) or "Unknown"
            last_error = f"The selected category is still '{selected_filter}'."
        except PlaywrightTimeoutError as error:
            last_error = str(error).splitlines()[0]
        except GameFilterError as error:
            last_error = str(error)
        dismiss_game_filter_dialog(page)
        if attempt < max_attempts:
            print(f"Game-category selection failed. Retrying ({attempt + 1}/{max_attempts})...")
            page.wait_for_timeout(700)
    raise GameFilterError(
        f"Could not select '{game_filter}' after {max_attempts} attempts: {last_error}"
    )


def get_search_box(page: Page) -> Locator:
    box = visible(page.get_by_placeholder(re.compile("\u5361\u540d.*\u7de8\u865f")))
    if box is None:
        raise NeedsManualInput("The Kapaipai search field was not found.")
    return box


def kapaipai_code_variants(card_code: str) -> list[str]:
    """Return source and Kapaipai display forms, e.g. PAC1-JP018/PAC1-018."""
    canonical = clean(card_code).upper()
    variants = [canonical]
    match = re.fullmatch(r"(.+)-JP(\d{3,4})", canonical)
    if match:
        displayed = f"{match.group(1)}-{match.group(2)}"
        if displayed not in variants:
            variants.append(displayed)
    return variants


def kapaipai_result_code_pattern(card_code: str) -> re.Pattern[str]:
    """Match normal and catalog-qualified Kapaipai display codes.

    Kapaipai sometimes inserts a catalog number between the set code and the
    card number, for example AGOV(1202)-JP002 for source code AGOV-JP002.
    """
    canonical = clean(card_code).upper()
    match = re.fullmatch(r"(.+)-JP(\d{3,4})", canonical)
    if not match:
        alternatives = "|".join(re.escape(value) for value in kapaipai_code_variants(canonical))
        return re.compile(rf"^(?:{alternatives})\s*(?:[\uff08(]\u7570\u5716[)\uff09])?$", re.I)
    set_code, number = match.groups()
    set_pattern = re.escape(set_code)
    catalog_qualifier = r"(?:[\uff08(][^()\uff08\uff09]+[)\uff09])?"
    return re.compile(
        rf"^{set_pattern}{catalog_qualifier}-(?:JP)?{re.escape(number)}"
        rf"\s*(?:[\uff08(]\u7570\u5716[)\uff09])?$",
        re.I,
    )


def deduplicate_search_results(
    candidates: Iterable[SearchResultCandidate],
) -> list[SearchResultCandidate]:
    """Collapse duplicate DOM labels for the same visible card version."""
    unique: dict[tuple[str, str, bool], SearchResultCandidate] = {}
    for candidate in candidates:
        key = (
            normalize_for_match(candidate.code_text),
            candidate.rarity.upper(),
            candidate.is_alt_art,
        )
        unique.setdefault(key, candidate)
    return list(unique.values())


def looks_like_card_code(value: str) -> bool:
    normalized = clean(value).upper()
    if CARD_CODE_RE.search(normalized):
        return True
    return bool(re.fullmatch(r"[A-Z0-9]{2,12}(?:[\uff08(][^()\uff08\uff09]+[)\uff09])?-JP\d{3,4}", normalized))


def search_result_candidates(page: Page, card_code: str) -> list[SearchResultCandidate]:
    code_pattern = kapaipai_result_code_pattern(card_code)
    code_nodes = page.get_by_text(code_pattern, exact=True)
    results: list[SearchResultCandidate] = []
    for index in range(code_nodes.count()):
        code_node = code_nodes.nth(index)
        try:
            if not code_node.is_visible():
                continue
            code_text = clean(code_node.inner_text())
            row = code_node.locator("xpath=..")
            rarity = ""
            for text in row.locator("span").all_inner_texts():
                candidate = clean(text).upper()
                if (
                    candidate != code_text.upper()
                    and not looks_like_card_code(candidate)
                    and RESULT_RARITY_RE.fullmatch(candidate)
                ):
                    rarity = candidate
                    break
            if rarity:
                results.append(SearchResultCandidate(
                    code_text=code_text,
                    rarity=rarity,
                    is_alt_art="\u7570\u5716" in normalize_for_match(code_text),
                    locator=code_node,
                ))
        except PlaywrightTimeoutError:
            continue
    return deduplicate_search_results(results)


def available_rarities(page: Page, card_code: str, is_alt_art: bool = False) -> list[str]:
    choices: list[str] = []
    for result in search_result_candidates(page, card_code):
        if result.is_alt_art == is_alt_art and result.rarity not in choices:
            choices.append(result.rarity)
    return choices


def choose_rarity(item: Listing, choices: list[str]) -> str:
    if item.rarity and item.rarity.upper() in choices:
        return item.rarity.upper()
    if not choices:
        raise CardNotFound("Kapaipai did not return a selectable card version.")
    if len(choices) == 1:
        only = choices[0]
        print(f"The rarity was ambiguous in the title, but {only} is the only result and was selected.")
        return only
    expected = item.rarity or item.rarity_hint or "Unknown"
    raise NeedsManualInput(
        f"Rarity '{expected}' does not uniquely match the search results: {', '.join(choices)}"
    )


def choose_search_result(item: Listing, candidates: list[SearchResultCandidate]) -> SearchResultCandidate:
    candidates = deduplicate_search_results(candidates)
    matching_version = [result for result in candidates if result.is_alt_art == item.is_alt_art]
    version_name = "alternate-art version" if item.is_alt_art else "standard version"
    if not matching_version:
        raise CardNotFound(f"Kapaipai did not find the {version_name} of {item.card_code}.")
    choices: list[str] = []
    for result in matching_version:
        if result.rarity not in choices:
            choices.append(result.rarity)
    rarity = choose_rarity(item, choices)
    matching_rarity = [result for result in matching_version if result.rarity == rarity]
    if len(matching_rarity) != 1:
        raise NeedsManualInput(
            f"The {version_name} rarity {rarity} for {item.card_code} is not unique."
        )
    return matching_rarity[0]


def search_and_open(page: Page, item: Listing, game_filter: str) -> str:
    if page.url != SEARCH_URL:
        page.goto(SEARCH_URL, wait_until="domcontentloaded")
    candidates: list[SearchResultCandidate] = []
    last_error = "The search result did not become ready."
    for search_attempt in range(1, SEARCH_MAX_ATTEMPTS + 1):
        candidates = []
        try:
            close_update_dialog(page)
            ensure_ygo_filter(page, game_filter, force_reselect=True)
            box = get_search_box(page)
            box.fill(item.card_code)
            search = visible(page.get_by_role("button", name=re.compile("\u641c\u5c0b")))
            if search is None:
                raise NeedsManualInput("The search button was not found.")
            search.click()
            deadline = time.monotonic() + 12
            empty_result_streak = 0
            while time.monotonic() < deadline:
                selected_filter = current_game_filter(page)
                if selected_filter and selected_filter != game_filter:
                    last_error = f"The category changed to '{selected_filter}' after searching."
                    break
                candidates = search_result_candidates(page, item.card_code)
                if candidates and selected_filter == game_filter:
                    break
                body = page.locator("body").inner_text()
                if "\u6c92\u6709\u76f8\u95dc\u7d50\u679c" in body and selected_filter == game_filter:
                    empty_result_streak += 1
                    if empty_result_streak >= 4:
                        raise CardNotFound(f"Kapaipai did not find {item.card_code}.")
                else:
                    empty_result_streak = 0
                page.wait_for_timeout(250)
            if candidates and current_game_filter(page) == game_filter:
                break
            if not candidates and current_game_filter(page) == game_filter:
                last_error = (
                    f"No selectable result for {item.card_code} was detected before the timeout."
                )
        except CardNotFound:
            raise
        except (GameFilterError, NeedsManualInput, PlaywrightTimeoutError) as error:
            last_error = str(error).splitlines()[0]
        if search_attempt < SEARCH_MAX_ATTEMPTS:
            print(
                f"Search attempt {search_attempt}/{SEARCH_MAX_ATTEMPTS} failed: {last_error} "
                "Retrying the category and search..."
            )
            dismiss_game_filter_dialog(page)
            page.wait_for_timeout(700)
    if not candidates:
        raise GameFilterError(
            f"The category and search could not be verified for {item.card_code} after "
            f"{SEARCH_MAX_ATTEMPTS} attempts: {last_error}"
        )
    selected = choose_search_result(item, candidates)
    rarity = selected.rarity
    selected.locator.click()
    try:
        page.get_by_role(
            "heading",
            name=re.compile(rf"{re.escape(selected.code_text)}\s+{re.escape(rarity)}", re.I),
        ).wait_for(timeout=12000)
    except PlaywrightTimeoutError:
        heading_text = page.locator("h1,h2,h3,h4").all_inner_texts()
        if not any(selected.code_text in text and rarity in text for text in heading_text):
            raise NeedsManualInput(
                "The card code, alternate-art state, or rarity could not be verified on the card page."
            )
    return rarity


def confirm_dialog(page: Page) -> bool:
    for label in ("\u78ba\u5b9a", "\u78ba\u8a8d", "\u5132\u5b58", "\u5b8c\u6210"):
        button = visible(page.get_by_role("button", name=label, exact=True))
        if button:
            button.click()
            page.wait_for_timeout(250)
            return True
    return False


def quick_panel(page: Page) -> Locator:
    add_button = visible(page.get_by_role("button", name="\u65b0\u589e\u5546\u54c1", exact=True))
    if add_button is None:
        add_button = visible(page.get_by_text("\u65b0\u589e\u5546\u54c1", exact=True))
    if add_button:
        selectors = (
            "xpath=ancestor::div[.//*[normalize-space()='\u65b0\u589e\u5099\u8a3b'] and .//*[contains(normalize-space(),'\u5eab\u5b58')]][1]",
            "xpath=ancestor::div[.//*[normalize-space()='\u65b0\u589e\u5099\u8a3b'] and .//*[normalize-space()='\u5728\u552e']][1]",
            "xpath=ancestor::div[.//*[normalize-space()='\u65b0\u589e\u5099\u8a3b']][1]",
        )
        for selector in selectors:
            panel = add_button.locator(selector)
            if panel.count():
                return panel
    return page.locator("body")


def fill_labelled_number(scope: Any, labels: Iterable[str], value: int) -> bool:
    label_re = re.compile("|".join(re.escape(label) for label in labels))
    candidates = [scope.get_by_label(label_re), scope.get_by_placeholder(label_re)]
    for locator in candidates:
        field = visible(locator)
        if field:
            field.fill(str(value))
            return True
    return False


def fill_quick_currency_input(page: Page, scope: Locator, value: int) -> bool:
    fields = scope.locator("input[inputmode='decimal'], input")
    for index in range(fields.count()):
        field = fields.nth(index)
        try:
            if not field.is_visible():
                continue
            raw = clean(field.input_value())
            if "$" not in raw:
                continue
            field.click()
            field.fill(str(value))
            field.press("Tab")
            page.wait_for_timeout(300)
            updated = clean(field.input_value())
            match = re.search(r"\d[\d,]*", updated)
            if match and int(match.group().replace(",", "")) == value:
                return True
        except (PlaywrightTimeoutError, ValueError):
            continue
    return False


def quick_currency_value(scope: Locator) -> int | None:
    fields = scope.locator("input[inputmode='decimal'], input")
    for index in range(fields.count()):
        field = fields.nth(index)
        try:
            if not field.is_visible():
                continue
            raw = clean(field.input_value())
            if "$" not in raw:
                continue
            match = re.search(r"\d[\d,]*", raw)
            if match:
                return int(match.group().replace(",", ""))
        except (PlaywrightTimeoutError, ValueError):
            continue
    try:
        current, _, _ = currency_stepper(scope)
        return current
    except NeedsManualInput:
        pass
    return None


def stepper_button(container: Locator, direction: int) -> Locator | None:
    if direction > 0:
        pattern = re.compile(r"^\s*\+\s*$|\u589e\u52a0|\u52a0\u4e00|increase", re.I)
    else:
        pattern = re.compile(r"^\s*[-\u2212\u2013\u2014]\s*$|\u6e1b\u5c11|\u6e1b\u4e00|decrease", re.I)
    button = visible(container.get_by_role("button", name=pattern))
    if button is not None:
        return button
    text_target = visible(container.get_by_text(pattern))
    if text_target is not None:
        return text_target
    clickables = container.locator("button,[role=button],[aria-label],[title]")
    for index in range(clickables.count()):
        candidate = clickables.nth(index)
        try:
            if not candidate.is_visible():
                continue
            signature = " ".join(
                clean(value)
                for value in (
                    candidate.inner_text(),
                    candidate.get_attribute("aria-label"),
                    candidate.get_attribute("title"),
                )
                if value
            )
            if pattern.search(signature):
                return candidate
        except PlaywrightTimeoutError:
            continue
    return None


def nearest_stepper_container(anchor: Locator, max_levels: int = 7) -> tuple[Locator, Locator, Locator] | None:
    container = anchor.locator("xpath=..")
    for _ in range(max_levels):
        if not container.count():
            return None
        plus = stepper_button(container, 1)
        minus = stepper_button(container, -1)
        if plus is not None and minus is not None:
            return container, plus, minus
        container = container.locator("xpath=..")
    return None


def control_number(container: Locator, currency: bool) -> int | None:
    inputs = container.locator("input")
    for index in range(inputs.count()):
        field = inputs.nth(index)
        try:
            if not field.is_visible():
                continue
            raw = clean(field.input_value())
            match = re.search(r"\d[\d,]*", raw)
            if match:
                return int(match.group().replace(",", ""))
        except (PlaywrightTimeoutError, ValueError):
            continue
    text = clean(container.inner_text())
    if currency:
        match = re.search(r"(?:NT)?\$\s*([0-9,]+)", text, re.I)
    else:
        match = re.search(r"(?:\u5728\u552e|\u5eab\u5b58)\D*([0-9,]+)", text, re.S)
        if match is None:
            match = re.search(r"(?<![$\d])([0-9][0-9,]*)(?!\d)", text)
    if not match:
        return None
    return int(match.group(1).replace(",", ""))


def currency_stepper(scope: Locator) -> tuple[int, Locator, Locator]:
    inputs = scope.locator("input")
    for index in range(inputs.count()):
        field = inputs.nth(index)
        try:
            if not field.is_visible():
                continue
            group = nearest_stepper_container(field)
            if group is None:
                continue
            container, plus, minus = group
            text = clean(container.inner_text())
            raw_value = clean(field.input_value())
            if "\u5728\u552e" in text and "$" not in text and "$" not in raw_value:
                continue
            current = control_number(container, currency=True)
            if current is not None:
                return current, plus, minus
        except PlaywrightTimeoutError:
            continue
    amounts = scope.get_by_text(re.compile(r"^\s*(?:NT)?\$\s*[0-9,]+\s*$", re.I))
    for index in range(amounts.count()):
        amount = amounts.nth(index)
        try:
            if not amount.is_visible():
                continue
            group = nearest_stepper_container(amount)
            if group is None:
                continue
            container, plus, minus = group
            current = control_number(container, currency=True)
            if current is not None:
                return current, plus, minus
        except PlaywrightTimeoutError:
            continue
    raise NeedsManualInput("The price stepper controls were not found.")


def quantity_stepper(scope: Locator) -> tuple[int, Locator, Locator]:
    labels = scope.get_by_text("\u5728\u552e", exact=True)
    for index in range(labels.count()):
        label = labels.nth(index)
        try:
            if not label.is_visible():
                continue
            group = nearest_stepper_container(label)
            if group is None:
                continue
            container, plus, minus = group
            current = control_number(container, currency=False)
            if current is not None:
                return current, plus, minus
        except PlaywrightTimeoutError:
            continue
    inputs = scope.locator("input")
    for index in range(inputs.count()):
        field = inputs.nth(index)
        try:
            if not field.is_visible():
                continue
            group = nearest_stepper_container(field)
            if group is None:
                continue
            container, plus, minus = group
            if "\u5728\u552e" not in clean(container.inner_text()):
                continue
            current = control_number(container, currency=False)
            if current is not None:
                return current, plus, minus
        except PlaywrightTimeoutError:
            continue
    raise NeedsManualInput("The listed-quantity stepper controls were not found.")


def wait_for_listing_controls(page: Page, timeout_seconds: int = 20) -> None:
    deadline = time.monotonic() + timeout_seconds
    last_error = ""
    while time.monotonic() < deadline:
        panel = quick_panel(page)
        try:
            currency_stepper(panel)
            quantity_stepper(panel)
            return
        except NeedsManualInput as error:
            last_error = str(error)
        page.wait_for_timeout(250)
    raise NeedsManualInput(f"Quick-listing controls timed out on the card page: {last_error}")


def adjust_stepper(
    page: Page,
    target: int,
    reader: Any,
    max_clicks: int,
    field_name: str,
) -> None:
    visits: dict[int, int] = {}
    for _ in range(max_clicks + 1):
        current, plus, minus = reader()
        if current == target:
            return
        visits[current] = visits.get(current, 0) + 1
        if visits[current] >= 3:
            raise NeedsManualInput(f"{field_name} could not be changed from {current} to {target}.")
        button = plus if current < target else minus
        if not button.is_enabled():
            raise NeedsManualInput(
                f"The {field_name} stepper is disabled; current value {current}, target {target}."
            )
        button.click()
        page.wait_for_timeout(90)
    current, _, _ = reader()
    raise NeedsManualInput(
        f"Changing {field_name} from {current} to {target} exceeded the {max_clicks}-click limit."
    )


def set_price(page: Page, price: int, max_clicks: int) -> None:
    panel = quick_panel(page)
    if fill_labelled_number(panel, ("\u50f9\u683c", "\u552e\u50f9"), price):
        page.wait_for_timeout(200)
        try:
            if currency_stepper(panel)[0] == price:
                return
        except NeedsManualInput:
            return
    if fill_quick_currency_input(page, panel, price):
        return
    adjust_stepper(page, price, lambda: currency_stepper(quick_panel(page)), max_clicks, "price")
    actual, _, _ = currency_stepper(quick_panel(page))
    if actual != price:
        raise NeedsManualInput(f"The price readback is {actual}, not the target {price}.")


def set_quantity(page: Page, quantity: int, max_clicks: int) -> None:
    panel = quick_panel(page)
    if fill_labelled_number(panel, ("\u6578\u91cf", "\u5eab\u5b58", "\u5728\u552e"), quantity):
        page.wait_for_timeout(200)
        try:
            if quantity_stepper(panel)[0] == quantity:
                return
        except NeedsManualInput:
            return
    adjust_stepper(page, quantity, lambda: quantity_stepper(quick_panel(page)), max_clicks, "quantity")
    actual, _, _ = quantity_stepper(quick_panel(page))
    if actual != quantity:
        raise NeedsManualInput(f"The quantity readback is {actual}, not the target {quantity}.")


def set_note(page: Page, note: str) -> None:
    panel = quick_panel(page)
    button = visible(panel.get_by_text("\u65b0\u589e\u5099\u8a3b", exact=True))
    if button is None:
        button = visible(panel.get_by_role("button", name="\u65b0\u589e\u5099\u8a3b", exact=True))
    if button is None:
        raise NeedsManualInput("The add-note button was not found.")
    button.click()
    page.wait_for_timeout(250)
    field = visible(page.locator("textarea"))
    if field is None:
        field = visible(page.get_by_placeholder(re.compile("\u5099\u8a3b|\u5206\u6578|\u8aaa\u660e")))
    if field is None:
        field = visible(page.get_by_role("textbox"))
    if field is None:
        raise NeedsManualInput("The note input field was not found after opening the note dialog.")
    field.fill(note)
    confirm_dialog(page)
    page.wait_for_timeout(250)
    if visible(quick_panel(page).get_by_text(note, exact=True)) is None:
        raise NeedsManualInput("The saved note could not be verified.")


def find_text_button(scope: Any, label: str) -> Locator | None:
    button = visible(scope.get_by_role("button", name=label, exact=True))
    if button is None:
        button = visible(scope.get_by_text(label, exact=True))
    return button


def wait_for_text_button(page: Page, label: str, timeout_seconds: int = 15) -> Locator:
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        button = find_text_button(page, label)
        if button is not None:
            return button
        page.wait_for_timeout(250)
    raise NeedsManualInput(f"The '{label}' control did not appear within {timeout_seconds} seconds.")


def wait_for_listing_action(page: Page, timeout_seconds: int = 15) -> tuple[str, Locator]:
    edit_label = "\u5b8c\u6574\u7de8\u8f2f"
    add_label = "\u65b0\u589e\u5546\u54c1"
    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        edit_button = find_text_button(page, edit_label)
        if edit_button is not None:
            return "edit", edit_button
        add_button = find_text_button(quick_panel(page), add_label)
        if add_button is None:
            add_button = find_text_button(page, add_label)
        if add_button is not None:
            return "add", add_button
        page.wait_for_timeout(250)
    raise NeedsManualInput(
        "Neither the full-edit control nor the add-product control appeared "
        f"within {timeout_seconds} seconds."
    )


def create_default_listing(
    page: Page,
    target_price: int | None = None,
    max_clicks: int = 1000,
) -> Locator:
    edit_label = "\u5b8c\u6574\u7de8\u8f2f"
    action, action_button = wait_for_listing_action(page)
    if action == "edit":
        print("An existing listing was detected. Reusing its full-edit action.")
        return action_button

    if target_price is not None:
        initial_price = quick_currency_value(quick_panel(page))
        if initial_price is None:
            print("The quick-listing price could not be identified. Leaving it unchanged.")
        elif initial_price <= 0:
            if target_price <= 0:
                raise NeedsManualInput("The product cannot be created with a zero price.")
            print(
                f"The quick-listing price is {initial_price}. "
                f"Setting it to {target_price} before product creation."
            )
            set_price(page, target_price, max_clicks)
            action, action_button = wait_for_listing_action(page)
            if action == "edit":
                print("An existing listing appeared after the price update. Reusing its full-edit action.")
                return action_button

    add_button = action_button

    click_error: PlaywrightTimeoutError | None = None
    try:
        add_button.click(timeout=5000)
    except PlaywrightTimeoutError as error:
        click_error = error
        print("The add-product control was replaced during the click. Verifying the resulting state.")

    try:
        edit_button = wait_for_text_button(page, edit_label)
    except NeedsManualInput as error:
        raise SubmissionUnverified(
            "The add-product control was clicked, but the full-edit action did not appear. "
            "The product may already exist, so the workflow stopped to prevent a duplicate."
        ) from (click_error or error)

    print("The default product was created and verified by the full-edit action.")
    return edit_button


def full_edit_panel(page: Page) -> Locator | None:
    selector = (
        "xpath=ancestor::*[.//*[contains(normalize-space(),'\u5546\u54c1\u50f9\u683c')] and "
        ".//*[contains(normalize-space(),'\u5728\u552e\u6578\u91cf')] and "
        ".//*[contains(normalize-space(),'\u5546\u54c1\u5099\u8a3b')] and "
        ".//*[normalize-space()='\u4fdd\u5b58\u8b8a\u66f4']][1]"
    )
    required_labels = ("\u5546\u54c1\u50f9\u683c", "\u5728\u552e\u6578\u91cf", "\u5546\u54c1\u5099\u8a3b")
    note_fields = page.get_by_placeholder(re.compile("\u8acb\u8f38\u5165\u5099\u8a3b"))
    for index in range(note_fields.count()):
        note_field = note_fields.nth(index)
        try:
            if not note_field.is_visible():
                continue
            panel = note_field.locator(selector)
            if panel.count() and panel.is_visible():
                text = clean(panel.inner_text())
                if all(label in text for label in required_labels):
                    return panel
        except PlaywrightTimeoutError:
            continue
    return None


def open_full_edit(page: Page, edit_button: Locator, timeout_seconds: int = 15) -> Locator:
    click_error: PlaywrightTimeoutError | None = None
    try:
        edit_button.click(timeout=5000)
    except PlaywrightTimeoutError as error:
        click_error = error
        print("The full-edit control was replaced during the click. Verifying the resulting state.")

    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        panel = full_edit_panel(page)
        if panel is not None:
            if click_error is not None:
                print("The full-edit panel appeared, so the click was successful.")
            return panel
        page.wait_for_timeout(250)
    error = NeedsManualInput(f"The full-edit panel did not open within {timeout_seconds} seconds.")
    if click_error is not None:
        raise error from click_error
    raise error


def input_after_label(panel: Locator, label_text: str) -> Locator:
    label_pattern = re.compile(rf"^\s*{re.escape(label_text)}\s*[:\uff1a]?\s*$")
    labels = panel.get_by_text(label_pattern)
    for index in range(labels.count()):
        label = labels.nth(index)
        try:
            if not label.is_visible():
                continue
            field = visible(label.locator("xpath=following::input[1]"))
            if field is not None:
                return field
            container = label.locator("xpath=ancestor::*[.//input][1]")
            field = visible(container.locator("input")) if container.count() else None
            if field is not None:
                return field
        except PlaywrightTimeoutError:
            continue
    raise NeedsManualInput(f"The input field for '{label_text}' was not found.")


def input_integer(field: Locator) -> int | None:
    try:
        match = re.search(r"\d[\d,]*", clean(field.input_value()))
        return int(match.group().replace(",", "")) if match else None
    except (PlaywrightTimeoutError, ValueError):
        return None


def set_full_edit_number(page: Page, panel: Locator, label_text: str, target: int) -> None:
    field = input_after_label(panel, label_text)
    try:
        field.click()
        field.fill(str(target))
        field.press("Tab")
        page.wait_for_timeout(300)
    except PlaywrightTimeoutError as error:
        raise NeedsManualInput(f"The '{label_text}' field could not be changed to {target}.") from error
    actual = input_integer(field)
    if actual != target:
        raise NeedsManualInput(f"The '{label_text}' readback is {actual}, not the target {target}.")


def set_full_edit_note(page: Page, panel: Locator, note: str) -> None:
    field = visible(panel.get_by_placeholder(re.compile("\u8acb\u8f38\u5165\u5099\u8a3b")))
    if field is None:
        field = visible(panel.locator("textarea"))
    if field is None:
        label = visible(panel.get_by_text(re.compile(r"^\s*\u5546\u54c1\u5099\u8a3b\s*[:\uff1a]?\s*$")))
        if label is not None:
            field = visible(label.locator("xpath=following::*[self::textarea or self::input][1]"))
    if field is None:
        raise NeedsManualInput("The product-note field was not found in the full-edit panel.")
    field.fill(note)
    field.press("Tab")
    page.wait_for_timeout(250)
    try:
        actual = clean(field.input_value())
    except PlaywrightTimeoutError as error:
        raise NeedsManualInput("The product note could not be read back.") from error
    if actual != clean(note):
        raise NeedsManualInput(f"The product-note readback is '{actual}', not the target note.")


def ensure_full_edit_listing_enabled(panel: Locator) -> None:
    switches = panel.get_by_role("switch")
    for index in range(switches.count()):
        switch = switches.nth(index)
        try:
            if not switch.is_visible():
                continue
            if clean(switch.get_attribute("aria-checked")).lower() == "false":
                switch.click()
            return
        except PlaywrightTimeoutError:
            continue
    checkboxes = panel.locator("input[type=checkbox]")
    for index in range(checkboxes.count()):
        checkbox = checkboxes.nth(index)
        try:
            if checkbox.is_visible() and not checkbox.is_checked():
                checkbox.click()
            return
        except PlaywrightTimeoutError:
            continue


def save_full_edit(page: Page, panel: Locator, timeout_seconds: int = 15) -> None:
    save_button = find_text_button(panel, "\u4fdd\u5b58\u8b8a\u66f4")
    if save_button is None:
        raise NeedsManualInput("The save-changes control was not found in the full-edit panel.")
    click_error: PlaywrightTimeoutError | None = None
    try:
        save_button.click(timeout=5000)
    except PlaywrightTimeoutError as error:
        click_error = error
        print("The save-changes control was replaced during the click. Verifying the resulting state.")

    deadline = time.monotonic() + timeout_seconds
    while time.monotonic() < deadline:
        if full_edit_panel(page) is None:
            if click_error is not None:
                print("The full-edit panel closed, so the changes were saved.")
            return
        page.wait_for_timeout(250)
    error = SubmissionUnverified(
        "Save changes was clicked, but the full-edit panel did not close. "
        "The workflow stopped because the update could not be verified."
    )
    if click_error is not None:
        raise error from click_error
    raise error


def submit_listing(page: Page, item: Listing, max_clicks: int) -> None:
    edit_button = create_default_listing(page, item.price, max_clicks)
    panel = open_full_edit(page, edit_button)
    set_full_edit_number(page, panel, "\u5546\u54c1\u50f9\u683c", item.price)
    set_full_edit_number(page, panel, "\u5728\u552e\u6578\u91cf", item.quantity)
    set_full_edit_note(page, panel, item.note)
    ensure_full_edit_listing_enabled(panel)
    save_full_edit(page, panel)


def progress_path(source: Path) -> Path:
    return source.with_name(f"{source.stem}_kapaipai_progress.json")


def diagnostic_control_inventory(page: Page) -> str:
    lines: list[str] = []
    edit_panel = full_edit_panel(page)
    scope = edit_panel if edit_panel is not None else quick_panel(page)
    elements = scope.locator("input,textarea,button,[role=button],[role=switch],[aria-label],[title],span,div")
    for index in range(min(elements.count(), 300)):
        element = elements.nth(index)
        try:
            if not element.is_visible():
                continue
            text = clean(element.inner_text())
            value = clean(element.get_attribute("value"))
            role = clean(element.get_attribute("role"))
            aria = clean(element.get_attribute("aria-label"))
            title = clean(element.get_attribute("title"))
            if not (
                text in {"+", "\uff0b", "-", "\u2212", "\u2013", "\u2014"}
                or value
                or role == "button"
                or aria
                or title
            ):
                continue
            html = clean(element.evaluate("element => element.outerHTML"))[:1000]
            lines.append(html)
        except Exception:
            continue
    return "\n".join(lines) or "<No relevant controls were found>"


def load_progress(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"completed_rows": [], "items": []}
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_progress(path: Path, progress: dict[str, Any]) -> None:
    temp = path.with_suffix(path.suffix + ".tmp")
    with temp.open("w", encoding="utf-8") as handle:
        json.dump(progress, handle, ensure_ascii=False, indent=2)
    temp.replace(path)


def save_diagnostic(page: Page, source: Path, item: Listing, stage: str, error: Exception) -> Path:
    root = source.parent / "kapaipai_diagnostics"
    root.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = root / f"{stamp}_row{item.row}_{item.card_code}_{stage}"
    try:
        edit_panel = full_edit_panel(page)
        panel_text = (edit_panel if edit_panel is not None else quick_panel(page)).inner_text()
    except Exception as panel_error:
        panel_text = f"<Could not read the quick-listing panel: {panel_error}>"
    try:
        body_text = page.locator("body").inner_text()[:30000]
    except Exception as body_error:
        body_text = f"<Could not read the page: {body_error}>"
    try:
        control_inventory = diagnostic_control_inventory(page)
    except Exception as inventory_error:
        control_inventory = f"<Could not inspect the controls: {inventory_error}>"
    detail = (
        f"Program version: {PROGRAM_VERSION}\n"
        f"Time: {datetime.now().isoformat(timespec='seconds')}\n"
        f"Stage: {stage}\n"
        f"Source row: {item.row}\n"
        f"Card code: {item.card_code}\n"
        f"Alternate art: {'Yes' if item.is_alt_art else 'No'}\n"
        f"Product: {item.name}\n"
        f"Target price: {item.price}\n"
        f"Target quantity: {item.quantity}\n"
        f"Target note: {item.note}\n"
        f"Error: {error}\n"
        f"URL: {page.url}\n\n"
        f"--- Active listing panel ---\n{panel_text}\n\n"
        f"--- Active listing control HTML ---\n{control_inventory}\n\n"
        f"--- Visible page text (first 30000 characters) ---\n{body_text}\n"
    )
    path = prefix.with_suffix(".txt")
    path.write_text(detail, encoding="utf-8")
    try:
        page.screenshot(path=str(prefix.with_suffix(".png")), full_page=True)
    except Exception:
        pass
    return path


def run_upload(source: Path, listings: list[Listing], config: dict[str, Any], args: argparse.Namespace) -> None:
    if sync_playwright is None:
        raise RuntimeError("Playwright is not installed. Run: py -m pip install -r requirements.txt")
    state_path = progress_path(source)
    progress = load_progress(state_path)
    completed = {int(row) for row in progress.get("completed_rows", [])}
    verified_rows = {
        int(record["row"])
        for record in progress.get("items", [])
        if record.get("verified_fields") is True and record.get("row") is not None
    }
    legacy_rows = sorted(completed - verified_rows)
    if legacy_rows:
        print(
            "Warning: legacy completed records do not contain price verification. "
            "These Excel rows will still be skipped to prevent duplicates: "
            + ", ".join(map(str, legacy_rows))
        )
        print("Check the prices of these existing products on Kapaipai before continuing.")
    queue = [item for item in listings if item.row not in completed and item.row >= args.start_row]
    if args.limit:
        queue = queue[: args.limit]
    print(
        f"Loaded {len(listings)} listings; {len(completed)} completed; "
        f"{len(queue)} queued for this run."
    )
    if not queue:
        print("There are no products waiting to be listed.")
        return
    if not args.yes:
        confirm = input("This will create products on Kapaipai. Type UPLOAD to continue: ").strip()
        if confirm != "UPLOAD":
            print("Canceled. No products were created.")
            return
    browser_process, debug_port = launch_manual_login_browser(source, args.browser_path)
    with sync_playwright() as playwright:
        browser = playwright.chromium.connect_over_cdp(f"http://127.0.0.1:{debug_port}")
        try:
            if not browser.contexts:
                raise RuntimeError("Connected to Chrome, but no usable browser window was found.")
            context = browser.contexts[0]
            page = select_kapaipai_page(context)
            wait_for_login(page)
            if not page.url.startswith(SEARCH_URL):
                page.goto(SEARCH_URL, wait_until="domcontentloaded")
            close_update_dialog(page)
            if not is_logged_in(page):
                wait_for_login(page)
            for position, item in enumerate(queue, 1):
                print(f"\n[{position}/{len(queue)}] Row {item.row} {item.card_code} | {item.name}")
                try:
                    rarity = search_and_open(page, item, config["game_filter"])
                    item.rarity = rarity
                    submit_listing(page, item, int(config.get("max_click_adjustment", 1000)))
                    item.status = "Listed"
                    item.message = f"{item.card_code} {rarity}"
                    completed.add(item.row)
                    progress["completed_rows"] = sorted(completed)
                    progress.setdefault("items", []).append({
                        **asdict(item),
                        "time": datetime.now().isoformat(timespec="seconds"),
                        "program_version": PROGRAM_VERSION,
                        "verified_fields": True,
                    })
                    save_progress(state_path, progress)
                    print(
                        f"Completed: {item.card_code} {rarity}, price {item.price}, "
                        f"quantity {item.quantity}."
                    )
                    page.wait_for_timeout(int(float(args.delay) * 1000))
                except GameFilterError as error:
                    item.status = "Game category failed"
                    item.message = str(error)
                    progress.setdefault("items", []).append({
                        **asdict(item),
                        "time": datetime.now().isoformat(timespec="seconds"),
                        "program_version": PROGRAM_VERSION,
                        "verified_fields": False,
                    })
                    save_progress(state_path, progress)
                    diagnostic = save_diagnostic(page, source, item, "game_category", error)
                    print(f"Game-category selection failed: {error}")
                    print(f"Diagnostic file: {diagnostic}")
                    print("The workflow stopped to prevent the batch from using the wrong category.")
                    raise
                except CardNotFound as error:
                    item.status = "Card not found"
                    item.message = str(error)
                    progress.setdefault("items", []).append({
                        **asdict(item),
                        "time": datetime.now().isoformat(timespec="seconds"),
                        "program_version": PROGRAM_VERSION,
                        "verified_fields": False,
                    })
                    save_progress(state_path, progress)
                    print(f"Skipped because the card was not found: {error}")
                    continue
                except NeedsManualInput as error:
                    item.status = "Upload stopped"
                    item.message = str(error)
                    progress.setdefault("items", []).append({
                        **asdict(item),
                        "time": datetime.now().isoformat(timespec="seconds"),
                        "program_version": PROGRAM_VERSION,
                        "verified_fields": False,
                    })
                    save_progress(state_path, progress)
                    diagnostic = save_diagnostic(page, source, item, "automatic_fill", error)
                    print(f"Upload stopped: {error}")
                    print(f"Diagnostic file: {diagnostic}")
                    print("The workflow stopped at the first failed product.")
                    raise
                except SubmissionUnverified as error:
                    item.status = "Submission requires verification"
                    item.message = str(error)
                    progress.setdefault("items", []).append({
                        **asdict(item),
                        "time": datetime.now().isoformat(timespec="seconds"),
                        "program_version": PROGRAM_VERSION,
                        "verified_fields": False,
                    })
                    save_progress(state_path, progress)
                    diagnostic = save_diagnostic(page, source, item, "submission_verification", error)
                    print(str(error))
                    print(f"Diagnostic file: {diagnostic}")
                    print("The workflow stopped to prevent duplicate listings.")
                    raise
                except Exception as error:
                    diagnostic = save_diagnostic(page, source, item, "unexpected_error", error)
                    print(f"Unexpected error: {error}")
                    print(f"Diagnostic file: {diagnostic}")
                    raise
        finally:
            browser.close()
            try:
                browser_process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                pass
    print(f"\nProgress file: {state_path}")


def resolve_input(raw: str | None) -> Path:
    if raw:
        path = Path(raw.strip().strip('"')).expanduser()
        if not path.exists():
            raise FileNotFoundError(f"Excel workbook not found: {path}")
        preview_suffix = "_kapaipai_preview"
        if path.stem.endswith(preview_suffix):
            original = path.with_name(f"{path.stem[:-len(preview_suffix)]}{path.suffix}")
            if not original.exists():
                raise FileNotFoundError(
                    "A Kapaipai preview workbook was selected, but its original Ruten workbook "
                    f"was not found in the same folder: {original.name}"
                )
            print(f"Preview workbook detected. Using the original Ruten workbook: {original.name}")
            path = original
        return path.resolve()
    preferred = list(Path.cwd().glob("ruten_products_*.xlsx"))
    fallback = list(Path.cwd().glob("*.xlsx"))
    candidates = sorted(
        {path for path in [*preferred, *fallback] if not path.stem.endswith("_kapaipai_preview")},
        key=lambda path: (path not in preferred, -path.stat().st_mtime),
    )
    if not candidates:
        raise FileNotFoundError("No Ruten workbook was found. Use --input to specify an Excel file.")
    return candidates[0].resolve()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Create Kapaipai Yu-Gi-Oh! listings from a Ruten Excel workbook."
    )
    parser.add_argument("--input", help="Ruten Excel path; defaults to the latest workbook")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="Rarity and consignor rules file")
    parser.add_argument("--execute", action="store_true", help="Create products; otherwise generate a preview")
    parser.add_argument("--start-row", type=int, default=1, help="Only process rows at or after this source row")
    parser.add_argument("--limit", type=int, default=0, help="Maximum products to list; 0 means unlimited")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay after each successful listing")
    parser.add_argument("--yes", action="store_true", help="Skip the UPLOAD confirmation prompt")
    parser.add_argument("--browser-path", default="", help="Path to Google Chrome or Microsoft Edge")
    parser.add_argument("--no-manual-fallback", action="store_true", help=argparse.SUPPRESS)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        source = resolve_input(args.input)
    except Exception as error:
        print(f"Workflow failed: {error}", file=sys.stderr)
        return 1

    log_path = source.with_name(f"{source.stem}_kapaipai_upload.log")
    with log_path.open("a", encoding="utf-8", newline="") as transcript:
        out = TeeTextIO(sys.stdout, transcript)
        err = TeeTextIO(sys.stderr, transcript)
        with contextlib.redirect_stdout(out), contextlib.redirect_stderr(err):
            try:
                print(f"\n===== {datetime.now().isoformat(timespec='seconds')} =====")
                print(f"Kapaipai uploader version: {PROGRAM_VERSION}")
                print(f"Log file: {log_path}")
                config = load_config(Path(args.config).resolve())
                listings, skipped = read_listings(source, config)
                if not listings:
                    raise RuntimeError("The Excel workbook contains no processable Yu-Gi-Oh! products.")
                if not args.execute:
                    preview = write_preview(source, listings, skipped)
                    unresolved = sum(not item.rarity for item in listings)
                    print(f"Preview created: {preview}")
                    print(
                        f"Processable: {len(listings)}; unresolved rarity: {unresolved}; "
                        f"skipped: {len(skipped)}."
                    )
                    print("Use --execute only after reviewing the preview.")
                    return 0
                run_upload(source, listings, config, args)
                return 0
            except KeyboardInterrupt:
                print("\nStopped by the user.")
                return 130
            except Exception as error:
                print(f"Workflow failed: {error}", file=sys.stderr)
                return 1


if __name__ == "__main__":
    raise SystemExit(main())

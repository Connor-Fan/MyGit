import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from ruten_exporter import (
    DEFAULT_PAGE_DELAY,
    Product,
    ProductCollector,
    build_parser,
    detail_quantity_from_html,
    export_excel,
    parse_price,
    parse_quantity,
    page_state_changed,
    products_from_json,
    store_list_url,
    turn_to_next_page,
    validate_store_url,
    wait_for_page_change,
)


ROOT = Path(__file__).resolve().parents[1]


class ParserTests(unittest.TestCase):
    def test_price_and_quantity_example(self):
        text = "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b DBWS-JP023 \u8ced\u4e0a\u9748\u9b42! (\u4eae\u9762) NT$80 \u51711\u4ef6"
        self.assertEqual(parse_price(text), 80)
        self.assertEqual(parse_quantity(text), 1)

    def test_json_product(self):
        payload = {
            "items": [{
                "itemId": "22612345678901",
                "itemName": "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u6e2c\u8a66\u5546\u54c1",
                "price": 80,
                "stockQuantity": 3,
            }]
        }
        products = products_from_json(payload, "test")
        self.assertEqual(len(products), 1)
        self.assertEqual(products[0].quantity, 3)
        self.assertEqual(products[0].price, 80)

    def test_detail_html(self):
        html = '<html><script type="application/ld+json">{"@type":"Product","name":"\u6e2c\u8a66\u5546\u54c1\u540d\u7a31","offers":{"price":80,"url":"https://www.ruten.com.tw/item/show?22612345678901","inventoryLevel":{"value":7}}}</script></html>'
        self.assertEqual(detail_quantity_from_html(html, "test"), 7)

    def test_store_root_is_changed_to_full_list(self):
        self.assertEqual(
            store_list_url("https://www.ruten.com.tw/store/qzecrvyn/"),
            "https://www.ruten.com.tw/store/qzecrvyn/list",
        )

    def test_store_url_is_required_and_validated(self):
        url = "https://www.ruten.com.tw/store/example_seller/"
        args = build_parser().parse_args(["--store-url", url, "--demo"])
        self.assertEqual(args.store_url, url)
        self.assertEqual(validate_store_url(f'"{url}"'), url)

    def test_page_delay_defaults_to_one_second(self):
        args = build_parser().parse_args([
            "--store-url", "https://www.ruten.com.tw/store/example_seller/", "--demo"
        ])
        self.assertEqual(args.page_delay, DEFAULT_PAGE_DELAY)
        self.assertEqual(args.page_delay, 1)

    def test_page_state_requires_changed_displayed_products(self):
        before = {
            "url": "https://www.ruten.com.tw/store/example_seller/list",
            "page_number": "13",
            "item_ids": ("A", "B"),
        }
        unchanged = dict(before)
        changed = dict(before, page_number="14", item_ids=("C", "D"))
        self.assertFalse(page_state_changed(before, unchanged))
        self.assertTrue(page_state_changed(before, changed))

    def test_wait_for_page_change_polls_until_item_ids_change(self):
        class FakePage:
            url = "https://www.ruten.com.tw/store/example_seller/list"

            def __init__(self):
                self.states = [
                    {"pageNumber": "13", "itemIds": ["A", "B"]},
                    {"pageNumber": "14", "itemIds": ["C", "D"]},
                ]

            def evaluate(self, _script):
                return self.states.pop(0)

            def wait_for_timeout(self, _milliseconds):
                return None

        before = {
            "url": FakePage.url,
            "page_number": "13",
            "item_ids": ("A", "B"),
        }
        changed = wait_for_page_change(FakePage(), before, timeout_seconds=0.1)
        self.assertIsNotNone(changed)
        self.assertEqual(changed["page_number"], "14")
        self.assertEqual(changed["item_ids"], ("C", "D"))

    def test_wait_for_page_change_accepts_new_api_products_when_dom_is_unchanged(self):
        class FakePage:
            url = "https://www.ruten.com.tw/store/example_seller/list"

            def evaluate(self, _script):
                return {"pageNumber": "1", "itemIds": []}

            def wait_for_timeout(self, _milliseconds):
                return None

        counts = iter([30, 60])
        changed = wait_for_page_change(
            FakePage(),
            {"url": FakePage.url, "page_number": "1", "item_ids": ()},
            timeout_seconds=0.1,
            product_count=lambda: next(counts),
            before_product_count=30,
        )
        self.assertIsNotNone(changed)
        self.assertEqual(changed["product_count"], 60)

    def test_successful_api_page_change_clicks_next_only_once(self):
        collector = ProductCollector()
        for index in range(30):
            collector.add(Product(f"Initial product {index}", 10, 1, item_id=f"A{index:08d}"))

        class FakePage:
            url = "https://www.ruten.com.tw/store/example_seller/list"

            def evaluate(self, script):
                if "activeSelectors" in script:
                    return {"pageNumber": "1", "itemIds": []}
                return []

            def wait_for_timeout(self, _milliseconds):
                return None

        class FakeNextPage:
            def __init__(self):
                self.click_count = 0

            def scroll_into_view_if_needed(self):
                return None

            def click(self):
                self.click_count += 1
                for index in range(30):
                    collector.add(Product(f"Next product {index}", 10, 1, item_id=f"B{index:08d}"))

        next_page = FakeNextPage()
        with patch("ruten_exporter.find_next_page", return_value=next_page):
            page_turned, next_available = turn_to_next_page(
                FakePage(),
                collector,
                {"url": FakePage.url, "page_number": "1", "item_ids": ()},
                page_delay=0,
            )

        self.assertTrue(page_turned)
        self.assertTrue(next_available)
        self.assertEqual(next_page.click_count, 1)
        self.assertEqual(len(collector.products), 60)

    def test_run_batch_prompts_for_store_url(self):
        batch = (ROOT / "run.bat").read_text(encoding="utf-8")
        self.assertIn("Enter or paste the Ruten store URL", batch)
        self.assertIn('--store-url "%STORE_URL%"', batch)
        self.assertNotIn('call :run_python "ruten_exporter.py"\n', batch)

    def test_run_batch_uses_user_facing_kapaipai_descriptions(self):
        batch = (ROOT / "run.bat").read_text(encoding="utf-8")
        uploader = (ROOT / "kapaipai_uploader.py").read_text(encoding="utf-8")
        self.assertIn("Create Kapaipai preview and skipped-items report", batch)
        self.assertIn("Upload products to Kapaipai", batch)
        self.assertNotIn("Use --execute only after reviewing the preview.", uploader)


class ExcelTests(unittest.TestCase):
    def test_excel_has_requested_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            export_excel([
                Product("\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b DBWS-JP023 \u8ced\u4e0a\u9748\u9b42! (\u4eae\u9762)", 80, 1)
            ], output, "https://www.ruten.com.tw/store/qzecrvyn/")
            workbook = load_workbook(output, data_only=False)
            sheet = workbook["Products"]
            self.assertEqual(
                [sheet.cell(1, col).value for col in range(1, 4)],
                ["Product Name", "Quantity", "Price"],
            )
            self.assertEqual(sheet["A2"].value, "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b DBWS-JP023 \u8ced\u4e0a\u9748\u9b42! (\u4eae\u9762)")
            self.assertEqual(sheet["B2"].value, 1)
            self.assertEqual(sheet["C2"].value, 80)


if __name__ == "__main__":
    unittest.main()

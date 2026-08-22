import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock, patch

from openpyxl import Workbook, load_workbook

from kapaipai_uploader import (
    CardNotFound,
    GameFilterError,
    Listing,
    ListingEditTarget,
    ManualListingRequired,
    NeedsManualInput,
    PlaywrightError,
    PlaywrightTimeoutError,
    SearchResultMismatch,
    SearchResultCandidate,
    adjust_stepper,
    build_note,
    choose_rarity,
    choose_search_result,
    click_filter_target,
    control_number,
    create_default_listing,
    custom_game_filter_trigger,
    deduplicate_search_results,
    detect_alt_art,
    displayed_search_result_codes,
    ensure_ygo_filter,
    ensure_inline_edit_note,
    ensure_inline_stock_capacity,
    is_logged_in,
    is_rarity_label,
    inline_decimal_field,
    kapaipai_result_code_pattern,
    looks_like_card_code,
    load_config,
    latest_skipped_records,
    missing_search_result_error,
    normalize_code,
    open_full_edit,
    parenthesized_rarity_text,
    parse_listing,
    quick_currency_value,
    read_listings,
    resolve_input,
    run_upload,
    save_inline_edit,
    save_full_edit,
    search_and_open,
    set_price,
    set_full_edit_number,
    stepper_button,
    submit_listing,
    wait_for_listing_action,
    wait_for_search_interface,
    wait_for_listing_controls,
    write_preview,
    write_skipped_report,
)


ROOT = Path(__file__).resolve().parents[1]


class ParserTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.config = load_config(ROOT / "kapaipai_rules.json")

    def test_sample_card(self):
        item = parse_listing(
            2,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b UT01-JP001 \u53e4\u4e4b\u79d8\u5100 (\u8d85\u6846.\u91d1\u4eae) \u5bc4\u8ce3:D",
            1,
            "NT$120",
            "https://www.ruten.com.tw/item/show?22632351536244",
            self.config,
        )
        self.assertEqual(item.card_code, "UT01-JP001")
        self.assertEqual(item.rarity, "UR-OF")
        self.assertEqual(item.quantity, 1)
        self.assertEqual(item.price, 120)
        self.assertEqual(item.consignor, "D")
        self.assertEqual(item.note, "95\uff5e97\u5206\uff5c\u5bc4\u8ce3:D")

    def test_d_consignor_is_not_younger_brother(self):
        item = parse_listing(
            2,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b RV01-JP003 \u90aa\u60e1\u7bb1 (\u4eae\u9762) \u5bc4\u8ce3:D",
            2,
            10,
            "",
            self.config,
        )
        self.assertEqual(item.consignor, "D")
        self.assertEqual(item.note, "95\uff5e97\u5206\uff5c\u5bc4\u8ce3:D")

    def test_note_rules(self):
        self.assertEqual(build_note("\u65e5\u7d19", "95\uff5e97\u5206", "\u9673"), "\u65e5\u7d19\uff5c95\uff5e97\u5206\uff5c\u5bc4\u8ce3:\u9673")
        self.assertEqual(build_note("", "95\uff5e97\u5206", "\u9673"), "95\uff5e97\u5206\uff5c\u5bc4\u8ce3:\u9673")
        self.assertEqual(build_note("", "95\uff5e97\u5206", ""), "95\uff5e97\u5206")

    def test_new_rarity_aliases(self):
        cases = {
            "\u666e\u947d": "NPR",
            "\u96b1\u666e": "NR",
            "\u91d1\u947d": "QCSER",
            "\u96d5\u947d": "CR",
            "\u7d05\u5b57\u534a\u947d": "SER-SRV",
            "\u96f7\u5c04": "HR",
            "\u7d05\u4eae": "UR",
            "\u85cd\u4eae": "UR",
            "\u85cd\u947d": "SER",
            "\u534a\u947d\u788e\u947d": "SEPR",
            "\u91d1\u4eae\u788e\u947d": "UPR",
            "\u4eae\u9762\u788e\u947d": "SPR",
            "\u534a\u947d\u5168\u947d": "SEPR",
            "\u91d1\u4eae\u5168\u947d": "UPR",
            "\u4eae\u9762\u5168\u947d": "SPR",
        }
        for hint, expected in cases.items():
            with self.subTest(hint=hint):
                item = parse_listing(
                    2,
                    f"\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b TEST-JP001 \u6e2c\u8a66\u5361 ({hint})",
                    1,
                    100,
                    "",
                    self.config,
                )
                self.assertEqual(item.rarity_hint, hint)
                self.assertEqual(item.rarity, expected)

    def test_config_rarity_alias_is_source_of_truth(self):
        config = dict(self.config)
        config["rarity_aliases"] = dict(self.config["rarity_aliases"])
        config["rarity_aliases"]["\u666e\u947d"] = ""
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "stale_rules.json"
            path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
            loaded = load_config(path)
            self.assertEqual(loaded["rarity_aliases"]["\u666e\u947d"], "")

    def test_config_consignor_alias_is_source_of_truth(self):
        config = dict(self.config)
        config["consignor_aliases"] = dict(self.config["consignor_aliases"])
        config["consignor_aliases"]["D"] = "\u5f1f"
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "stale_rules.json"
            path.write_text(json.dumps(config, ensure_ascii=False), encoding="utf-8")
            loaded = load_config(path)
            self.assertEqual(loaded["consignor_aliases"]["D"], "\u5f1f")

    def test_preview_input_resolves_to_original_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            original = root / "ruten_products_20260811.xlsx"
            preview = root / "ruten_products_20260811_kapaipai_preview.xlsx"
            original.touch()
            preview.touch()
            self.assertEqual(resolve_input(str(preview)), original.resolve())

    def test_skipped_report_input_resolves_to_original_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            original = root / "ruten_products_20260811.xlsx"
            report = root / "ruten_products_20260811_kapaipai_skipped.xlsx"
            original.touch()
            report.touch()
            self.assertEqual(resolve_input(str(report)), original.resolve())

    def test_kapaipai_catalog_qualified_card_code(self):
        pattern = kapaipai_result_code_pattern("AGOV-JP002")
        self.assertRegex("AGOV-JP002", pattern)
        self.assertRegex("AGOV-002", pattern)
        self.assertRegex("AGOV(1202)-JP002", pattern)
        self.assertRegex("AGOV\uff081202\uff09-JP002", pattern)
        self.assertNotRegex("AGOV(1202)-JP003", pattern)

    def test_extended_yugioh_card_code_formats(self):
        codes = [
            "BLZD-JPS06",
            "BPRO-JPS07",
            "20CP-JPC04",
            "15AX-JPY52",
            "DS14-JPL28",
            "NCF1-JPP01",
            "DUEA-JA045",
            "EE3-147",
        ]
        for code in codes:
            with self.subTest(code=code):
                self.assertEqual(normalize_code(f"\u904a\u6232\u738b {code} test"), code)
                self.assertTrue(looks_like_card_code(code))
                self.assertRegex(code, kapaipai_result_code_pattern(code))

    def test_kapaipai_card_code_accepts_named_artwork_variants(self):
        pattern = kapaipai_result_code_pattern("SD47-JP001")
        self.assertRegex("SD47-JP001(\u7570\u5716A)", pattern)
        self.assertRegex("SD47-JP001(\u7570\u5716B)", pattern)
        self.assertRegex("SD47-JP001\uff08\u7570\u5716C\uff09", pattern)

    def test_search_history_card_codes_are_not_rarity_labels(self):
        self.assertTrue(looks_like_card_code("DAMA-JP008"))
        self.assertTrue(looks_like_card_code("AGOV(1202)-JP002"))
        self.assertFalse(looks_like_card_code("UR"))
        self.assertFalse(looks_like_card_code("SER-SRV"))

    def test_numeric_prefix_rarities_are_supported(self):
        self.assertTrue(is_rarity_label("20SER"))
        self.assertTrue(is_rarity_label("25SER"))
        self.assertTrue(is_rarity_label("SER-SRV"))
        self.assertFalse(is_rarity_label("06498706"))
        self.assertFalse(is_rarity_label("DAMA-JP008"))

        item = Listing(1112, "test", 3, 3000, "ETCO-JP071", "", "20SER", "", "", "note")
        candidates = [
            SearchResultCandidate("ETCO(1012)-JP071", "20SER", False, object()),
            SearchResultCandidate("ETCO(1012)-JP071", "SER", False, object()),
            SearchResultCandidate("ETCO(1012)-JP071", "SR", False, object()),
        ]
        self.assertEqual(choose_search_result(item, candidates).rarity, "20SER")

    def test_duplicate_dom_card_version_is_collapsed(self):
        first_locator = object()
        duplicate_locator = object()
        candidates = [
            SearchResultCandidate("CF01-JP026", "UR", False, first_locator),
            SearchResultCandidate("CF01-JP026", "UR", False, duplicate_locator),
        ]
        unique = deduplicate_search_results(candidates)
        self.assertEqual(len(unique), 1)
        self.assertIs(unique[0].locator, first_locator)

        item = Listing(415, "test", 1, 20, "CF01-JP026", "", "UR", "", "", "note")
        selected = choose_search_result(item, candidates)
        self.assertIs(selected.locator, first_locator)

    def test_screenshot_titles_and_invisible_characters(self):
        titles = [
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011 \u904a\u6232\u738b DBCB-JP036 \u767b\u9f8d\u83ef\u8f49\u751f\u7d0b (\u666e\u947d)",
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011 \u904a\u6232\u738b DBCB-JP037 \u767b\u9f8d\u83ef\u6050\u6727\u9580\uff08\u666e\u947d\uff09",
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011 \u904a\u6232\u738b DBJH-JP012 \u92d2\u5229\u5c0f\u9b3c \u934a\u5b50 (\u666e\u200b\u947d)",
        ]
        for title in titles:
            with self.subTest(title=title):
                item = parse_listing(2, title, 1, 5, "", self.config)
                self.assertEqual(item.rarity_hint, "\u666e\u947d")
                self.assertEqual(item.rarity, "NPR")

    def test_longer_rarity_alias_still_wins(self):
        item = parse_listing(
            2,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b TEST-JP001 \u6e2c\u8a66\u5361 (\u8d85\u6846.\u767d\u947d)",
            1,
            100,
            "",
            self.config,
        )
        self.assertEqual(item.rarity_hint, "\u8d85\u6846.\u767d\u947d")
        self.assertEqual(item.rarity, "PSER-OF")

    def test_rarity_is_only_read_inside_parentheses(self):
        cases = [
            (
                "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b ALIN-JP077 \u9ad8\u8cb4\u9ec3\u91d1\u9109\u7684\u9ec3\u91d1\u570b\u85e5 (\u9280\u5b57)",
                "\u9280\u5b57",
                "R",
            ),
            (
                "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b LEDE-JP020 \u9ec3\u91d1\u9a55\u50b2 \u8def\u9014\u9a55\u50b2 (\u666e\u5361)",
                "\u666e\u5361",
                "N",
            ),
            (
                "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b LEDE-JP041 \u9ec3\u91d1\u9a55\u50b2 \u6bb2\u6ec5\u8005\uff08\u4eae\u9762\uff09",
                "\u4eae\u9762",
                "SR",
            ),
            (
                "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b GS02-JP016 \u6280\u80fd\u525d\u596a \u6280\u80fd\u62bd\u53d6 (\u9ec3\u91d1)",
                "\u9ec3\u91d1",
                "GR",
            ),
            (
                "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b ETCO-JP071 \u878d\u5408\u6d3e\u5175 (\u7d05\u947d)",
                "\u7d05\u947d",
                "20SER",
            ),
        ]
        for title, expected_hint, expected_rarity in cases:
            with self.subTest(title=title):
                item = parse_listing(2, title, 1, 10, "", self.config)
                self.assertEqual(item.rarity_hint, expected_hint)
                self.assertEqual(item.rarity, expected_rarity)

    def test_rarity_words_outside_parentheses_are_ignored(self):
        title = "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b TEST-JP001 \u9ec3\u91d1\u5168\u947d\u7d05\u947d\u6e2c\u8a66\u5361"
        item = parse_listing(2, title, 1, 100, "", self.config)
        self.assertEqual(parenthesized_rarity_text(title), "")
        self.assertEqual(item.rarity_hint, "")
        self.assertEqual(item.rarity, "")

    def test_multiple_parentheses_are_combined(self):
        title = "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b TEST-JP001 \u6e2c\u8a66\u5361 (\u4e8c\u624b)\uff08\u7d05\u4eae\uff09"
        self.assertEqual(parenthesized_rarity_text(title), "\u4e8c\u624b \u7d05\u4eae")
        item = parse_listing(2, title, 1, 100, "", self.config)
        self.assertEqual(item.rarity_hint, "\u7d05\u4eae")
        self.assertEqual(item.rarity, "UR")

    def test_alt_art_is_detected_only_inside_parentheses(self):
        alt = parse_listing(
            1374,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b PAC1-JP018 \u511a\u7121\u6c34 \u511a\u7121\u6728 (\u7570\u5716.\u534a\u947d)",
            1,
            50,
            "",
            self.config,
        )
        normal = parse_listing(
            1375,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b PAC1-JP018 \u7570\u5716\u4ecb\u7d39 \u511a\u7121\u6c34\u6728 (\u534a\u947d)",
            1,
            20,
            "",
            self.config,
        )
        self.assertTrue(detect_alt_art(alt.name))
        self.assertTrue(alt.is_alt_art)
        self.assertEqual(alt.rarity, "SER")
        self.assertFalse(normal.is_alt_art)

    def test_search_result_matches_alt_art_before_rarity(self):
        candidates = [
            SearchResultCandidate("PAC1-018(\u7570\u5716)", "PSER", True),
            SearchResultCandidate("PAC1-018(\u7570\u5716)", "SER", True),
            SearchResultCandidate("PAC1-018", "PSER", False),
            SearchResultCandidate("PAC1-018", "SER", False),
            SearchResultCandidate("PAC1-018", "NPR", False),
        ]
        alt = Listing(2, "alt", 1, 50, "PAC1-JP018", "\u534a\u947d", "SER", "", "", "95\uff5e97\u5206", True)
        normal = Listing(3, "normal", 1, 20, "PAC1-JP018", "\u534a\u947d", "SER", "", "", "95\uff5e97\u5206")
        self.assertEqual(choose_search_result(alt, candidates).code_text, "PAC1-018(\u7570\u5716)")
        self.assertEqual(choose_search_result(normal, candidates).code_text, "PAC1-018")

    def test_mismatched_visible_results_are_skipped_instead_of_category_failure(self):
        page = Mock()
        page.locator.return_value.inner_text.return_value = "\n".join([
            "\u641c\u5c0b",
            "TW03-052",
            "20AP-JP087",
            "DUEA(901)-JP053",
            "BLZD-JPS17",
            "EE3-147",
            "TW03-054",
        ])
        item = Listing(
            3036,
            "\u904a\u6232\u738b TW03-JP053 \u5967\u7c73\u52a0\u661f\u8056 (\u666e\u947d)",
            2,
            5,
            "TW03-JP053",
            "\u666e\u947d",
            "NPR",
            "",
            "",
            "95\uff5e97\u5206",
        )

        self.assertEqual(
            displayed_search_result_codes(page),
            [
                "TW03-052",
                "20AP-JP087",
                "DUEA(901)-JP053",
                "BLZD-JPS17",
                "EE3-147",
                "TW03-054",
            ],
        )
        with patch("kapaipai_uploader.current_game_filter", return_value="\u904a\u6232\u738b\u65e5\u6587"):
            error = missing_search_result_error(
                page,
                item,
                "\u904a\u6232\u738b\u65e5\u6587",
                "No selectable result",
            )
        self.assertIsInstance(error, SearchResultMismatch)
        self.assertIn("TW03-JP053", str(error))
        self.assertIn("TW03-052", str(error))

    def test_missing_result_still_stops_when_game_category_is_wrong(self):
        page = Mock()
        item = Listing(3036, "test", 2, 5, "TW03-JP053", "", "NPR", "", "", "note")
        with patch("kapaipai_uploader.current_game_filter", return_value="\u5bf6\u53ef\u5922\u7e41\u4e2d"):
            error = missing_search_result_error(
                page,
                item,
                "\u904a\u6232\u738b\u65e5\u6587",
                "The category changed",
            )
        self.assertIsInstance(error, GameFilterError)

    def test_incomplete_search_ui_is_reloaded_before_retry(self):
        page = Mock()
        page.url = "https://trade.kapaipai.tw/search"
        search_button = Mock()
        search_box = Mock()
        selected = SearchResultCandidate("WPP4-JP050", "N", False, Mock())
        item = Listing(
            3167,
            "\u904a\u6232\u738b WPP4-JP050 \u632f\u5b50\u7279\u6025\u50be\u659c\u5feb\u8eca (\u666e\u5361)",
            4,
            5,
            "WPP4-JP050",
            "\u666e\u5361",
            "N",
            "",
            "",
            "95\uff5e97\u5206",
        )

        with (
            patch("kapaipai_uploader.close_update_dialog"),
            patch(
                "kapaipai_uploader.ensure_ygo_filter",
                side_effect=[GameFilterError("The game filter did not load within 30 seconds after login."), None],
            ) as ensure_filter,
            patch("kapaipai_uploader.get_search_box", return_value=search_box),
            patch("kapaipai_uploader.visible", return_value=search_button),
            patch("kapaipai_uploader.current_game_filter", return_value="\u904a\u6232\u738b\u65e5\u6587"),
            patch("kapaipai_uploader.search_result_candidates", return_value=[selected]),
            patch("kapaipai_uploader.dismiss_game_filter_dialog"),
        ):
            rarity = search_and_open(page, item, "\u904a\u6232\u738b\u65e5\u6587")

        self.assertEqual(rarity, "N")
        self.assertEqual(ensure_filter.call_count, 2)
        page.goto.assert_called_once_with(
            "https://trade.kapaipai.tw/search",
            wait_until="domcontentloaded",
            timeout=20000,
        )
        page.wait_for_timeout.assert_any_call(1500)

    def test_same_set_multiple_artworks_require_manual_listing(self):
        item = Listing(
            2528,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b SD47-JP001 \u9752\u773c\u767d\u9f8d EX\u7248 (\u666e\u5361)",
            1,
            1,
            "SD47-JP001",
            "\u666e\u5361",
            "N",
            "",
            "",
            "95\uff5e97\u5206",
        )
        candidates = [
            SearchResultCandidate("SD47-JP001(\u7570\u5716A)", "N", True),
            SearchResultCandidate("SD47-JP001(\u7570\u5716B)", "N", True),
            SearchResultCandidate("SD47-JP001(\u7570\u5716C)", "N", True),
        ]

        with self.assertRaisesRegex(ManualListingRequired, "multiple artwork variants"):
            choose_search_result(item, candidates)

    def test_single_same_set_result_ignores_cross_set_alt_art_label(self):
        item = Listing(
            1658,
            "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b LOCH-JP030 \u9ed1\u9b54\u5973 (\u7570\u5716.\u91d1\u4eae)",
            1,
            20,
            "LOCH-JP030",
            "\u91d1\u4eae",
            "SER",
            "",
            "",
            "95\uff5e97\u5206",
            True,
        )
        candidate = SearchResultCandidate("LOCH-JP030", "SER", False, object())

        self.assertIs(choose_search_result(item, [candidate]), candidate)

    def test_preview_has_alt_art_column(self):
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "source.xlsx"
            source.touch()
            item = parse_listing(
                1374,
                "\u3010\u7f85\u52c3\u5c0f\u8216\u3011\u904a\u6232\u738b PAC1-JP018 \u511a\u7121\u6c34\u6728 (\u7570\u5716.\u534a\u947d)",
                1,
                50,
                "",
                self.config,
            )
            preview = write_preview(source, [item])
            workbook = load_workbook(preview, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Upload Preview"])
            sheet = workbook["Upload Preview"]
            headers = [cell.value for cell in sheet[1]]
            self.assertIn("Alternate Art", headers)
            alt_cell = sheet.cell(2, headers.index("Alternate Art") + 1)
            self.assertEqual(alt_cell.value, "Yes")
            self.assertEqual(alt_cell.fill.fgColor.rgb, "00FFC7CE")
            workbook.close()

    def test_skipped_report_uses_latest_status_per_row(self):
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "ruten_products_20260820.xlsx"
            preview_skipped = [
                {
                    "row": 6,
                    "name": "Invalid source product",
                    "reason": "A Yu-Gi-Oh! card code was not found.",
                    "quantity": 1,
                    "price": 20,
                    "source_url": "https://www.ruten.com.tw/item/show?invalid",
                }
            ]
            progress = {
                "items": [
                    {
                        "row": 2,
                        "card_code": "OLD-JP001",
                        "name": "Previously missing",
                        "status": "Card not found",
                        "message": "Missing before retry",
                        "quantity": 1,
                        "price": 5,
                        "source_url": "https://www.ruten.com.tw/item/show?old",
                        "time": "2026-08-22T01:00:00",
                    },
                    {
                        "row": 2,
                        "card_code": "OLD-JP001",
                        "name": "Previously missing",
                        "status": "Listed",
                        "message": "Listed after retry",
                        "quantity": 1,
                        "price": 5,
                        "source_url": "https://www.ruten.com.tw/item/show?old",
                        "time": "2026-08-22T01:05:00",
                    },
                    {
                        "row": 3,
                        "card_code": "SD47-JP001",
                        "name": "Multiple artworks",
                        "status": "Manual listing required",
                        "message": "Multiple artwork variants",
                        "quantity": 1,
                        "price": 5,
                        "source_url": "https://www.ruten.com.tw/item/show?manual",
                        "time": "2026-08-22T01:10:00",
                    },
                    {
                        "row": 4,
                        "card_code": "TW03-JP053",
                        "name": "Mismatched result",
                        "status": "Search result mismatch",
                        "message": "Displayed another card",
                        "quantity": 2,
                        "price": 5,
                        "source_url": "https://www.ruten.com.tw/item/show?mismatch",
                        "time": "2026-08-22T01:15:00",
                    },
                    {
                        "row": 5,
                        "card_code": "MISS-JP001",
                        "name": "Missing card",
                        "status": "Card not found",
                        "message": "Kapaipai did not find the card",
                        "quantity": 1,
                        "price": 10,
                        "source_url": "https://www.ruten.com.tw/item/show?missing",
                        "time": "2026-08-22T01:20:00",
                    },
                ]
            }

            records = latest_skipped_records(progress, preview_skipped)
            self.assertEqual([record["row"] for record in records], [3, 4, 5, 6])

            output = write_skipped_report(source, progress, preview_skipped)
            self.assertEqual(output.name, "ruten_products_20260820_kapaipai_skipped.xlsx")
            workbook = load_workbook(output)
            sheet = workbook["Skipped Items"]
            self.assertEqual(sheet.max_row, 5)
            self.assertEqual(
                [sheet.cell(row, 1).value for row in range(2, 6)],
                [3, 4, 5, 6],
            )
            self.assertEqual(sheet.cell(2, 4).value, "Manual listing required")
            self.assertEqual(sheet.cell(3, 4).value, "Search result mismatch")
            self.assertEqual(sheet.cell(4, 4).value, "Card not found")
            self.assertEqual(sheet.cell(5, 4).value, "Preview validation failed")
            self.assertEqual(sheet.cell(2, 5).value, "Multiple artworks were found. Please list manually.")
            self.assertEqual(sheet.cell(3, 5).value, "The displayed card does not match the product. Skipped.")
            self.assertEqual(sheet.cell(4, 5).value, "No matching card was found on Kapaipai.")
            self.assertEqual(
                sheet.cell(3, 3).hyperlink.target,
                "https://www.ruten.com.tw/item/show?mismatch",
            )
            workbook.close()

    def test_rarity_selection_is_automatic(self):
        item = Listing(2, "test", 1, 100, "TEST-JP001", "", "", "", "", "95\uff5e97\u5206")
        self.assertEqual(choose_rarity(item, ["NPR"]), "NPR")
        with self.assertRaises(NeedsManualInput):
            choose_rarity(item, ["NPR", "SR"])

    def test_explicit_rarity_mismatch_is_card_not_found(self):
        item = Listing(3, "test", 3, 15, "18SP-JP203", "\u666e\u5361", "N", "", "", "95\uff5e97\u5206")
        candidates = [SearchResultCandidate("18SP-JP203", "SR", False, object())]

        with self.assertRaisesRegex(CardNotFound, "did not find rarity N.*available rarities: SR"):
            choose_search_result(item, candidates)

    def test_price_stepper_uses_observed_values(self):
        state = {"value": 10}

        class Button:
            def __init__(self, delta):
                self.delta = delta

            def is_enabled(self):
                return True

            def click(self):
                state["value"] += self.delta

        class Page:
            @staticmethod
            def wait_for_timeout(_milliseconds):
                return None

        plus, minus = Button(5), Button(-5)
        reader = lambda: (state["value"], plus, minus)
        adjust_stepper(Page(), 150, reader, 1000, "\u50f9\u683c")
        self.assertEqual(state["value"], 150)
        adjust_stepper(Page(), 5, reader, 1000, "\u50f9\u683c")
        self.assertEqual(state["value"], 5)

    def test_price_uses_direct_currency_input_before_stepper(self):
        page = object()
        panel = object()
        with (
            patch("kapaipai_uploader.quick_panel", return_value=panel),
            patch("kapaipai_uploader.fill_labelled_number", return_value=False),
            patch("kapaipai_uploader.fill_quick_currency_input", return_value=True) as direct_fill,
            patch("kapaipai_uploader.adjust_stepper") as adjust,
        ):
            set_price(page, 150, 1000)
        direct_fill.assert_called_once_with(page, panel, 150)
        adjust.assert_not_called()

    def test_control_number_can_read_input_value(self):
        class Field:
            @staticmethod
            def is_visible():
                return True

            @staticmethod
            def input_value():
                return "$1,250"

        class Inputs:
            @staticmethod
            def count():
                return 1

            @staticmethod
            def nth(_index):
                return Field()

        class Container:
            @staticmethod
            def locator(selector):
                self.assertEqual(selector, "input")
                return Inputs()

            @staticmethod
            def inner_text():
                return "\u2212 \uff0b"

        self.assertEqual(control_number(Container(), currency=True), 1250)

    def test_stepper_button_accepts_clickable_span_text(self):
        class Empty:
            @staticmethod
            def count():
                return 0

        class Span:
            @staticmethod
            def count():
                return 1

            def nth(self, _index):
                return self

            @staticmethod
            def is_visible():
                return True

        span = Span()

        class Container:
            @staticmethod
            def get_by_role(_role, **_kwargs):
                return Empty()

            @staticmethod
            def get_by_text(_pattern):
                return span

            @staticmethod
            def locator(_selector):
                return Empty()

        self.assertIs(stepper_button(Container(), 1), span)

    def test_game_filter_retries_and_clicks_visible_selected_item(self):
        state = {
            "selected": "\u5bf6\u53ef\u5922\u7e41\u4e2d",
            "opened": False,
            "dialog_open": False,
            "trigger_clicks": 0,
        }

        class Locator:
            def __init__(self, kind):
                self.kind = kind

            def count(self):
                return 0 if self.kind == "missing" else 1

            def nth(self, _index):
                if self.kind == "custom_list":
                    return Locator("custom_trigger")
                return self

            def is_visible(self):
                if self.kind == "dialog":
                    return state["dialog_open"]
                return True

            def inner_text(self):
                if self.kind == "dialog":
                    return "\u904a\u6232\u7be9\u9078"
                return state["selected"]

            def wait_for(self, **_kwargs):
                return None

            def get_by_text(self, text, **_kwargs):
                if self.kind == "dialog" and text == "\u904a\u6232\u738b\u65e5\u6587":
                    return Locator("option")
                return Locator("other")

            def locator(self, selector):
                if selector == "xpath=ancestor::*[@role='dialog']":
                    return Locator("missing")
                return Locator("missing")

            def click(self, **_kwargs):
                if self.kind == "custom_trigger":
                    state["trigger_clicks"] += 1
                    if state["trigger_clicks"] <= 2:
                        raise PlaywrightTimeoutError("\u6a21\u64ec\u9078\u55ae\u52d5\u756b\u5c1a\u672a\u7a69\u5b9a")
                    state["opened"] = True
                    state["dialog_open"] = True
                if self.kind == "option":
                    state["selected"] = "\u904a\u6232\u738b\u65e5\u6587"
                    state["dialog_open"] = False

        class Page:
            @staticmethod
            def locator(selector):
                if selector == "div[style*='cursor: pointer']":
                    return Locator("custom_list")
                return Locator("other")

            @staticmethod
            def get_by_role(role, **kwargs):
                if role == "dialog":
                    return Locator("dialog")
                if role == "option" and kwargs.get("name") == "\u904a\u6232\u738b\u65e5\u6587":
                    return Locator("option")
                return Locator("missing")

            @staticmethod
            def wait_for_timeout(_milliseconds):
                return None

        ensure_ygo_filter(Page(), "\u904a\u6232\u738b\u65e5\u6587")
        self.assertTrue(state["opened"])
        self.assertGreaterEqual(state["trigger_clicks"], 3)
        self.assertEqual(state["selected"], "\u904a\u6232\u738b\u65e5\u6587")

        previous_clicks = state["trigger_clicks"]
        state["opened"] = False
        ensure_ygo_filter(Page(), "\u904a\u6232\u738b\u65e5\u6587", force_reselect=True)
        self.assertTrue(state["opened"])
        self.assertGreater(state["trigger_clicks"], previous_clicks)

    def test_filter_option_outside_viewport_uses_dom_click(self):
        state = {"clicks": 0, "scrolled": 0, "dom_clicked": 0}

        class Target:
            @staticmethod
            def click(**_kwargs):
                state["clicks"] += 1
                raise PlaywrightError("Element is outside of the viewport")

            @staticmethod
            def scroll_into_view_if_needed(**_kwargs):
                state["scrolled"] += 1

            @staticmethod
            def evaluate(_script):
                state["dom_clicked"] += 1

        click_filter_target(Target())
        self.assertEqual(state["clicks"], 2)
        self.assertEqual(state["scrolled"], 1)
        self.assertEqual(state["dom_clicked"], 1)

    def test_filter_verification_ignores_options_inside_open_dialog(self):
        candidates = [
            ("\u904a\u6232\u738b\u65e5\u6587", True),
            ("\u5bf6\u53ef\u5922\u7e41\u4e2d", False),
        ]

        class CountLocator:
            def __init__(self, count):
                self._count = count

            def count(self):
                return self._count

        class Candidate:
            def __init__(self, label, inside_dialog):
                self.label = label
                self.inside_dialog = inside_dialog

            def is_visible(self):
                return True

            def inner_text(self):
                return self.label

            def locator(self, _selector):
                return CountLocator(1 if self.inside_dialog else 0)

        class CandidateList:
            def count(self):
                return len(candidates)

            def nth(self, index):
                return Candidate(*candidates[index])

        class Page:
            @staticmethod
            def locator(_selector):
                return CandidateList()

        _trigger, label = custom_game_filter_trigger(Page())
        self.assertEqual(label, "\u5bf6\u53ef\u5922\u7e41\u4e2d")

    def test_blank_page_is_not_mistaken_for_logged_in(self):
        class EmptyLocator:
            @staticmethod
            def count():
                return 0

            @staticmethod
            def inner_text():
                return ""

        class Page:
            url = "https://trade.kapaipai.tw/search"

            @staticmethod
            def locator(_selector):
                return EmptyLocator()

            @staticmethod
            def get_by_role(_role, **_kwargs):
                return EmptyLocator()

        self.assertFalse(is_logged_in(Page()))

    def test_search_interface_waits_for_delayed_controls_and_dialog(self):
        state = {"ticks": 0, "dialog_closed": False}

        class Locator:
            def __init__(self, kind):
                self.kind = kind

            def count(self):
                if self.kind == "custom_list":
                    return 1 if state["ticks"] >= 3 else 0
                if self.kind == "update_button":
                    return 1 if state["ticks"] >= 1 and not state["dialog_closed"] else 0
                return 0

            def nth(self, _index):
                if self.kind == "custom_list":
                    return Locator("custom_trigger")
                return self

            def is_visible(self):
                return True

            def inner_text(self):
                return "\u5bf6\u53ef\u5922\u7e41\u4e2d" if self.kind == "custom_trigger" else ""

            def click(self, **_kwargs):
                if self.kind == "update_button":
                    state["dialog_closed"] = True

            def locator(self, _selector):
                return Locator("missing")

        class Page:
            @staticmethod
            def locator(selector):
                if selector == "div[style*='cursor: pointer']":
                    return Locator("custom_list")
                return Locator("missing")

            @staticmethod
            def get_by_role(role, **kwargs):
                if role == "button" and kwargs.get("name") == "\u6211\u77e5\u9053\u4e86":
                    return Locator("update_button")
                return Locator("missing")

            @staticmethod
            def wait_for_timeout(_milliseconds):
                state["ticks"] += 1

        wait_for_search_interface(Page(), timeout_seconds=2)
        self.assertGreaterEqual(state["ticks"], 3)
        self.assertTrue(state["dialog_closed"])

    def test_listing_controls_wait_until_price_and_quantity_are_ready(self):
        state = {"ticks": 0}

        class Page:
            @staticmethod
            def wait_for_timeout(_milliseconds):
                state["ticks"] += 1

        def currency(_panel):
            if state["ticks"] < 2:
                raise NeedsManualInput("\u50f9\u683c\u63a7\u5236\u9805\u5c1a\u672a\u8f09\u5165")
            return 10, object(), object()

        def quantity(_panel):
            return 1, object(), object()

        with (
            patch("kapaipai_uploader.quick_panel", return_value=object()),
            patch("kapaipai_uploader.currency_stepper", side_effect=currency),
            patch("kapaipai_uploader.quantity_stepper", side_effect=quantity),
        ):
            wait_for_listing_controls(Page(), timeout_seconds=1)
        self.assertGreaterEqual(state["ticks"], 2)

    def test_submit_listing_uses_create_then_full_edit_flow(self):
        order = []
        edit_button = object()
        panel = object()
        item = Listing(2, "test", 3, 150, "TEST-JP001", "", "NPR", "", "", "note")

        with (
            patch(
                "kapaipai_uploader.create_default_listing",
                side_effect=lambda _page, _price, _max_clicks: order.append("create")
                or ListingEditTarget("full", edit_button),
            ),
            patch("kapaipai_uploader.inline_edit_panel", return_value=object()) as inline_probe,
            patch("kapaipai_uploader.open_full_edit", side_effect=lambda _page, _button: order.append("open_edit") or panel),
            patch("kapaipai_uploader.set_full_edit_number", side_effect=lambda _page, _panel, label, _value: order.append(label)),
            patch("kapaipai_uploader.set_full_edit_note", side_effect=lambda _page, _panel, _note: order.append("note")),
            patch("kapaipai_uploader.ensure_full_edit_listing_enabled", side_effect=lambda _panel: order.append("enabled")),
            patch("kapaipai_uploader.save_full_edit", side_effect=lambda _page, _panel: order.append("save")),
        ):
            submit_listing(object(), item, 1000)

        self.assertEqual(
            order,
            ["create", "open_edit", "\u5546\u54c1\u50f9\u683c", "\u5728\u552e\u6578\u91cf", "note", "enabled", "save"],
        )
        inline_probe.assert_not_called()

    def test_submit_listing_reuses_inline_editor(self):
        order = []
        panel = object()
        item = Listing(1627, "test", 1, 100, "QCAC-JP037", "", "QCSER", "", "", "95\uff5e97\u5206")

        with (
            patch("kapaipai_uploader.create_default_listing", return_value=ListingEditTarget("inline", panel)),
            patch(
                "kapaipai_uploader.ensure_inline_stock_capacity",
                side_effect=lambda _page, _panel, _quantity: order.append("stock"),
            ),
            patch(
                "kapaipai_uploader.set_inline_edit_number",
                side_effect=lambda _page, _panel, currency, _value: order.append("price" if currency else "quantity"),
            ),
            patch("kapaipai_uploader.ensure_inline_edit_note", side_effect=lambda _page, _panel, _note: order.append("note")),
            patch(
                "kapaipai_uploader.save_inline_edit",
                side_effect=lambda _page, _panel, *_expected: order.append("save"),
            ),
            patch("kapaipai_uploader.open_full_edit") as open_edit,
        ):
            submit_listing(object(), item, 1000)

        self.assertEqual(order, ["stock", "price", "quantity", "note", "save"])
        open_edit.assert_not_called()

    def test_existing_listing_reuses_full_edit_without_creating_duplicate(self):
        page = object()
        edit_button = object()
        with (
            patch("kapaipai_uploader.wait_for_listing_action", return_value=("edit", edit_button)),
            patch("kapaipai_uploader.quick_panel") as panel,
        ):
            target = create_default_listing(page)
        self.assertEqual(target.mode, "full")
        self.assertIs(target.locator, edit_button)
        panel.assert_not_called()

    def test_listing_action_waits_for_delayed_full_edit(self):
        state = {"ticks": 0}
        page = Mock()
        page.wait_for_timeout.side_effect = lambda _milliseconds: state.__setitem__("ticks", state["ticks"] + 1)
        edit_button = object()

        def find(_scope, label):
            if label == "\u5b8c\u6574\u7de8\u8f2f" and state["ticks"] >= 2:
                return edit_button
            return None

        with (
            patch("kapaipai_uploader.find_text_button", side_effect=find),
            patch("kapaipai_uploader.quick_panel", return_value=object()),
        ):
            action, button = wait_for_listing_action(page, timeout_seconds=1)
        self.assertEqual(action, "edit")
        self.assertIs(button, edit_button)
        self.assertGreaterEqual(state["ticks"], 2)

    def test_new_listing_clicks_add_product_then_waits_for_full_edit(self):
        page = object()
        quick_scope = object()
        add_button = Mock()
        edit_button = object()
        with (
            patch("kapaipai_uploader.quick_panel", return_value=quick_scope),
            patch("kapaipai_uploader.wait_for_listing_action", return_value=("add", add_button)),
            patch("kapaipai_uploader.wait_for_text_button", return_value=edit_button) as waiter,
        ):
            target = create_default_listing(page)
        self.assertEqual(target.mode, "full")
        self.assertIs(target.locator, edit_button)
        add_button.click.assert_called_once_with(timeout=5000)
        waiter.assert_called_once_with(page, "\u5b8c\u6574\u7de8\u8f2f")

    def test_zero_quick_price_is_initialized_before_add_product(self):
        page = object()
        quick_scope = object()
        add_button = Mock()
        edit_button = object()
        with (
            patch("kapaipai_uploader.quick_panel", return_value=quick_scope),
            patch("kapaipai_uploader.quick_currency_value", return_value=0),
            patch("kapaipai_uploader.set_price") as set_price,
            patch("kapaipai_uploader.wait_for_listing_action", return_value=("add", add_button)),
            patch("kapaipai_uploader.wait_for_text_button", return_value=edit_button),
        ):
            target = create_default_listing(page, 150, 1000)
        self.assertEqual(target.mode, "full")
        self.assertIs(target.locator, edit_button)
        set_price.assert_called_once_with(page, 150, 1000)
        add_button.click.assert_called_once_with(timeout=5000)

    def test_zero_price_update_can_open_existing_inline_editor(self):
        page = object()
        quick_scope = object()
        add_button = Mock()
        inline_panel = object()
        with (
            patch("kapaipai_uploader.quick_panel", return_value=quick_scope),
            patch("kapaipai_uploader.quick_currency_value", return_value=0),
            patch("kapaipai_uploader.set_price") as set_price,
            patch(
                "kapaipai_uploader.wait_for_listing_action",
                side_effect=[("add", add_button), ("inline_edit", inline_panel)],
            ),
        ):
            target = create_default_listing(page, 100, 1000)
        self.assertEqual(target.mode, "inline")
        self.assertIs(target.locator, inline_panel)
        set_price.assert_called_once_with(page, 100, 1000)
        add_button.click.assert_not_called()

    def test_zero_quick_price_can_be_read_from_stepper_text(self):
        class EmptyFields:
            @staticmethod
            def count():
                return 0

        class Scope:
            @staticmethod
            def locator(_selector):
                return EmptyFields()

        with patch("kapaipai_uploader.currency_stepper", return_value=(0, object(), object())):
            self.assertEqual(quick_currency_value(Scope()), 0)

    def test_unidentified_quick_price_does_not_block_add_product(self):
        page = object()
        quick_scope = object()
        add_button = Mock()
        edit_button = object()
        with (
            patch("kapaipai_uploader.quick_panel", return_value=quick_scope),
            patch("kapaipai_uploader.quick_currency_value", return_value=None),
            patch("kapaipai_uploader.set_price") as set_price,
            patch("kapaipai_uploader.wait_for_listing_action", return_value=("add", add_button)),
            patch("kapaipai_uploader.wait_for_text_button", return_value=edit_button),
        ):
            target = create_default_listing(page, 5, 1000)
        self.assertEqual(target.mode, "full")
        self.assertIs(target.locator, edit_button)
        set_price.assert_not_called()
        add_button.click.assert_called_once_with(timeout=5000)

    def test_add_product_timeout_is_accepted_when_full_edit_appears(self):
        page = object()
        quick_scope = object()
        add_button = Mock()
        add_button.click.side_effect = PlaywrightTimeoutError("detached after click")
        edit_button = object()
        with (
            patch("kapaipai_uploader.quick_panel", return_value=quick_scope),
            patch("kapaipai_uploader.wait_for_listing_action", return_value=("add", add_button)),
            patch("kapaipai_uploader.wait_for_text_button", return_value=edit_button),
        ):
            target = create_default_listing(page)
        self.assertEqual(target.mode, "full")
        self.assertIs(target.locator, edit_button)

    def test_full_edit_timeout_is_accepted_when_panel_appears(self):
        page = Mock()
        edit_button = Mock()
        edit_button.click.side_effect = PlaywrightTimeoutError("detached after click")
        panel = object()
        with patch("kapaipai_uploader.full_edit_panel", return_value=panel):
            self.assertIs(open_full_edit(page, edit_button), panel)
        edit_button.click.assert_called_once_with(timeout=5000)

    def test_save_timeout_is_accepted_when_panel_closes(self):
        page = Mock()
        panel = object()
        save_button = Mock()
        save_button.click.side_effect = PlaywrightTimeoutError("detached after click")
        with (
            patch("kapaipai_uploader.find_text_button", return_value=save_button),
            patch("kapaipai_uploader.full_edit_panel", return_value=None),
        ):
            save_full_edit(page, panel)
        save_button.click.assert_called_once_with(timeout=5000)

    def test_inline_decimal_fields_distinguish_price_and_quantity(self):
        price = Mock()
        price.is_visible.return_value = True
        price.input_value.return_value = "$100"
        quantity = Mock()
        quantity.is_visible.return_value = True
        quantity.input_value.return_value = "1"
        fields = Mock()
        fields.count.return_value = 2
        fields.nth.side_effect = [price, quantity, price, quantity]
        panel = Mock()
        panel.locator.return_value = fields

        self.assertIs(inline_decimal_field(panel, True), price)
        self.assertIs(inline_decimal_field(panel, False), quantity)

    def test_inline_stock_is_increased_before_listing_more_units(self):
        stock = Mock()
        stock.is_visible.return_value = True
        stock.input_value.side_effect = ["1", "3"]
        fields = Mock()
        fields.count.return_value = 1
        fields.nth.return_value = stock
        panel = Mock()
        panel.locator.return_value = fields
        page = Mock()

        ensure_inline_stock_capacity(page, panel, 3)

        stock.fill.assert_called_once_with("3")
        stock.press.assert_called_once_with("Tab")

    def test_inline_note_uses_add_note_control(self):
        page = object()
        panel = Mock()
        panel.inner_text.return_value = "\u65b0\u589e\u5099\u8a3b"
        with (
            patch("kapaipai_uploader.find_text_button", return_value=object()),
            patch("kapaipai_uploader.set_note") as set_note,
        ):
            ensure_inline_edit_note(page, panel, "95\uff5e97\u5206")
        set_note.assert_called_once_with(page, "95\uff5e97\u5206")

    def test_inline_save_timeout_is_accepted_when_editor_closes(self):
        page = Mock()
        panel = object()
        save_button = Mock()
        save_button.click.side_effect = PlaywrightTimeoutError("detached after click")
        with (
            patch("kapaipai_uploader.find_text_button", return_value=save_button),
            patch("kapaipai_uploader.inline_edit_panel", return_value=None),
        ):
            save_inline_edit(page, panel)
        save_button.click.assert_called_once_with(timeout=5000)

    def test_inline_save_accepts_matching_values_when_editor_stays_open(self):
        page = Mock()
        panel = object()
        save_button = Mock()
        with (
            patch("kapaipai_uploader.find_text_button", return_value=save_button),
            patch("kapaipai_uploader.inline_edit_panel", return_value=panel),
            patch("kapaipai_uploader.inline_edit_values_match", return_value=True),
            patch("kapaipai_uploader.control_is_inactive", side_effect=[False, True]),
        ):
            save_inline_edit(page, panel, 5, 2, "95\uff5e97\u5206")

        save_button.click.assert_called_once_with(timeout=5000)

    def test_full_edit_number_is_filled_and_verified(self):
        field = Mock()
        field.input_value.return_value = "150"
        page = Mock()
        with patch("kapaipai_uploader.input_after_label", return_value=field):
            set_full_edit_number(page, object(), "\u5546\u54c1\u50f9\u683c", 150)
        field.fill.assert_called_once_with("150")
        field.press.assert_called_once_with("Tab")

    def test_run_upload_stops_on_the_first_failed_product(self):
        source = Path("test_source.xlsx")
        listings = [
            Listing(2, "first", 1, 10, "TEST-JP001", "", "NPR", "", "", "note"),
            Listing(3, "second", 1, 20, "TEST-JP002", "", "NPR", "", "", "note"),
        ]
        args = SimpleNamespace(start_row=1, limit=0, yes=True, browser_path="", delay=0)
        page = SimpleNamespace(url="https://trade.kapaipai.tw/search")
        process = Mock()

        class Browser:
            contexts = [object()]

            @staticmethod
            def close():
                return None

        browser = Browser()

        class Chromium:
            @staticmethod
            def connect_over_cdp(_endpoint):
                return browser

        class PlaywrightContext:
            def __enter__(self):
                return SimpleNamespace(chromium=Chromium())

            def __exit__(self, *_args):
                return False

        with (
            patch("kapaipai_uploader.sync_playwright", return_value=PlaywrightContext()),
            patch("kapaipai_uploader.launch_manual_login_browser", return_value=(process, 1234)),
            patch("kapaipai_uploader.select_kapaipai_page", return_value=page),
            patch("kapaipai_uploader.wait_for_login"),
            patch("kapaipai_uploader.close_update_dialog"),
            patch("kapaipai_uploader.is_logged_in", return_value=True),
            patch("kapaipai_uploader.load_progress", return_value={"completed_rows": [], "items": []}),
            patch("kapaipai_uploader.save_progress"),
            patch("kapaipai_uploader.refresh_skipped_report", return_value=Path("skipped.xlsx")),
            patch("kapaipai_uploader.search_and_open", return_value="NPR") as search,
            patch("kapaipai_uploader.submit_listing", side_effect=NeedsManualInput("simulated failure")),
            patch("kapaipai_uploader.save_diagnostic", return_value=Path("diagnostic.txt")),
        ):
            with self.assertRaises(NeedsManualInput):
                run_upload(source, listings, {"game_filter": "\u904a\u6232\u738b\u65e5\u6587"}, args)
        self.assertEqual(search.call_count, 1)

    def test_run_upload_continues_when_card_is_not_found(self):
        source = Path("test_source.xlsx")
        listings = [
            Listing(2, "missing", 1, 10, "MISS-JP001", "", "NPR", "", "", "note"),
            Listing(3, "found", 1, 20, "TEST-JP002", "", "NPR", "", "", "note"),
        ]
        args = SimpleNamespace(start_row=1, limit=0, yes=True, browser_path="", delay=0)
        page = SimpleNamespace(url="https://trade.kapaipai.tw/search", wait_for_timeout=Mock())
        process = Mock()

        class Browser:
            contexts = [object()]

            @staticmethod
            def close():
                return None

        browser = Browser()

        class Chromium:
            @staticmethod
            def connect_over_cdp(_endpoint):
                return browser

        class PlaywrightContext:
            def __enter__(self):
                return SimpleNamespace(chromium=Chromium())

            def __exit__(self, *_args):
                return False

        with (
            patch("kapaipai_uploader.sync_playwright", return_value=PlaywrightContext()),
            patch("kapaipai_uploader.launch_manual_login_browser", return_value=(process, 1234)),
            patch("kapaipai_uploader.select_kapaipai_page", return_value=page),
            patch("kapaipai_uploader.wait_for_login"),
            patch("kapaipai_uploader.close_update_dialog"),
            patch("kapaipai_uploader.is_logged_in", return_value=True),
            patch("kapaipai_uploader.load_progress", return_value={"completed_rows": [], "items": []}),
            patch("kapaipai_uploader.save_progress"),
            patch("kapaipai_uploader.refresh_skipped_report", return_value=Path("skipped.xlsx")),
            patch(
                "kapaipai_uploader.search_and_open",
                side_effect=[CardNotFound("missing card"), "NPR"],
            ) as search,
            patch("kapaipai_uploader.submit_listing") as submit,
        ):
            run_upload(source, listings, {"game_filter": "\u904a\u6232\u738b\u65e5\u6587"}, args)
        self.assertEqual(search.call_count, 2)
        submit.assert_called_once()

    def test_read_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "\u5546\u54c1\u6e05\u55ae"
            sheet.append(["\u5546\u54c1\u540d\u7a31", "\u6578\u91cf", "\u50f9\u683c"])
            sheet.append(["\u904a\u6232\u738b UT01-JP001 \u53e4\u4e4b\u79d8\u5100 (\u8d85\u6846\u91d1\u4eae) \u65e5\u7d19 \u5bc4\u8ce3:\u9673", 1, 120])
            workbook.save(path)
            listings, skipped = read_listings(path, self.config)
            self.assertEqual(len(listings), 1)
            self.assertFalse(skipped)
            self.assertEqual(listings[0].note, "\u65e5\u7d19\uff5c95\uff5e97\u5206\uff5c\u5bc4\u8ce3:\u9673")

    def test_read_workbook_accepts_hyphenated_japanese_codes_only(self):
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Products"
            sheet.append(["Product Name", "Quantity", "Price"])
            sheet.append(["\u904a\u6232\u738b RB-60 \u5fc3\u8b8a (\u91d1\u4eae)", 1, 30])
            sheet.append(["\u904a\u6232\u738b QCTB \u7cbe\u7f8e\u9396\u5319\u6263 \u9470\u5319\u5708", 1, 10])
            sheet.append(["\u904a\u6232\u738b VOL2-37 \u7121\u6a19 \u9b54\u88dd\u9a0e\u58eb (\u666e\u5361)", 1, 5])
            sheet.append(["\u904a\u6232\u738b SDRB-AEP09 \u6708\u83ef\u9f8d (\u4eae\u9762)", 1, 10])
            sheet.append(["\u904a\u6232\u738b PHRA-EN099 \u96fb\u8166\u754c\u59ec (\u4eae\u9762)", 1, 10])
            workbook.save(path)

            listings, skipped = read_listings(path, self.config)

            self.assertEqual([item.card_code for item in listings], ["RB-60"])
            self.assertEqual(len(skipped), 4)
            self.assertEqual(skipped[0]["reason"], "No set-number card code was found. This may be merchandise.")
            self.assertEqual(skipped[1]["reason"], "Unnumbered cards are not supported. Please handle manually.")
            self.assertEqual(skipped[2]["reason"], "Asian-English cards are not supported. Please handle manually.")
            self.assertEqual(skipped[3]["reason"], "American-English cards are not supported. Please handle manually.")


if __name__ == "__main__":
    unittest.main()
